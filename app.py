# ============================================================
# IMPORTAÇÕES DO SISTEMA
# ============================================================

import os
import secrets
import smtplib
import pytz
import oracledb

from flask_wtf.csrf import CSRFProtect
from urllib.parse import quote
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO

from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from openpyxl import Workbook

from db import conectar_oracle


# ============================================================
# CONFIGURAÇÕES GERAIS DO SISTEMA
# ============================================================

# Define o diretório base do projeto.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Carrega as variáveis do arquivo .env em ambiente local.
# No Render, essas variáveis vêm do painel Environment.
load_dotenv(os.path.join(BASE_DIR, ".env"))

# ============================================================
# CONFIGURAÇÕES DE E-MAIL / SMTP
# ============================================================

# Servidor SMTP usado para envio de e-mails.
SMTP_HOST = os.getenv("SMTP_HOST")

# Porta SMTP. Por padrão, Gmail usa 587 com TLS.
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))

# Usuário da conta de e-mail remetente.
SMTP_USER = os.getenv("SMTP_USER")

# Senha de app do e-mail remetente.
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")

# E-mail que aparecerá como remetente.
MAIL_FROM = os.getenv("MAIL_FROM")

# URL base do sistema.
# Local: http://127.0.0.1:5000
# Produção: https://cabos-eleitorais.onrender.com
BASE_URL = os.getenv("BASE_URL", "http://127.0.0.1:5000")


# ============================================================
# CONFIGURAÇÕES ADMINISTRATIVAS
# ============================================================

# Usuário e senha administrativa opcional via variável de ambiente.
ADMIN_USER = os.getenv("ADMIN_USER")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")


# ============================================================
# INICIALIZAÇÃO DO FLASK
# ============================================================

# Cria a aplicação Flask.
app = Flask(__name__)
# ============================================================
# Expiração de sessão automática
# ============================================================

csrf = CSRFProtect(app)

app.permanent_session_lifetime = timedelta(minutes=30)

# Define a chave secreta usada para sessão, flash messages e tokens.
app.secret_key = os.getenv("FLASK_SECRET_KEY", "chave_padrao")

# Configurações de segurança da sessão.
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

# No Render/produção com HTTPS, deixe True.
# Em ambiente local sem HTTPS, pode causar problema de sessão.
if os.getenv("RENDER"):
    app.config["SESSION_COOKIE_SECURE"] = True

# Serializer usado para gerar tokens seguros, por exemplo recuperação de senha.
serializer = URLSafeTimedSerializer(app.secret_key)


# ============================================================
# DECORATOR: EXIGE LOGIN
# ============================================================

def login_required(f):
    """
    Decorator usado para proteger rotas que exigem usuário logado.
    Se não houver sessão ativa, redireciona para a tela de login.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("usuario_logado"):
            flash("Faça login para acessar o sistema.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


# ============================================================
# FILTRO JINJA: DATA/HORA NO HORÁRIO DE BRASÍLIA
# ============================================================

@app.template_filter("data_br")
def data_br(data):
    """
    Converte data/hora UTC para o horário de Brasília.
    Usado nos templates com: {{ item[5]|data_br }}
    """
    if not data:
        return "-"

    fuso_brasilia = pytz.timezone("America/Sao_Paulo")

    if data.tzinfo is None:
        data = pytz.utc.localize(data)

    data_brasilia = data.astimezone(fuso_brasilia)

    return data_brasilia.strftime("%d/%m/%Y %H:%M")


# ============================================================
# DECORATOR: EXIGE PERFIL DE ACESSO
# ============================================================

def perfil_required(*tipos_permitidos):
    """
    Decorator usado para restringir rotas por tipo de acesso.
    Exemplo:
    @perfil_required("ADMIN", "CHEFE_GABINETE")
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get("usuario_logado"):
                flash("Faça login para acessar o sistema.", "warning")
                return redirect(url_for("login"))

            tipo_usuario = session.get("tipo_acesso")

            if tipo_usuario not in tipos_permitidos:
                flash("Você não tem permissão para acessar esta área.", "danger")
                return redirect(url_for("home"))

            return f(*args, **kwargs)

        return decorated_function

    return decorator


# ============================================================
# HELPER: CONVERTER DATA DO FORMULÁRIO
# ============================================================

def converter_data(data_str):
    """
    Converte data recebida do formulário HTML.
    Entrada esperada: YYYY-MM-DD
    Retorno: objeto date ou None.
    """
    if not data_str:
        return None

    try:
        return datetime.strptime(data_str, "%Y-%m-%d").date()
    except ValueError:
        return None


# ============================================================
# HELPER: GERAR TOKEN DE CONVITE
# ============================================================

def gerar_token_convite():
    """
    Gera um token seguro para links de convite.
    """
    return secrets.token_urlsafe(32)


# ============================================================
# FUNÇÃO: ENVIO DE E-MAIL DE CONVITE
# ============================================================

def enviar_email_convite(destinatario, nome_cabo, link):
    """
    Envia e-mail de convite para cadastro.
    Retorna True quando o envio foi aceito pelo SMTP.
    Retorna False quando ocorre erro no envio.
    """
    try:
        assunto = "Sistema Lideranças - Convite de Cadastro"

        corpo_texto = f"""
Olá!

Você recebeu um convite para cadastro no Sistema de Lideranças.

Liderança responsável: {nome_cabo}

Acesse o link abaixo para concluir seu cadastro:
{link}

Se você não reconhece este convite, apenas ignore esta mensagem.

Atenciosamente,
Sistema de Lideranças
"""

        corpo_html = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
</head>
<body style="margin:0; padding:0; background-color:#f4f6f8; font-family:Arial, Helvetica, sans-serif;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background-color:#f4f6f8; padding:30px 0;">
        <tr>
            <td align="center">
                <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff; border-radius:10px; overflow:hidden; border:1px solid #e5e7eb;">
                    <tr>
                        <td style="background:#198754; padding:22px; text-align:center; color:#ffffff;">
                            <h1 style="margin:0; font-size:24px;">Sistema de Lideranças</h1>
                        </td>
                    </tr>

                    <tr>
                        <td style="padding:30px; color:#111827;">
                            <h2 style="margin-top:0; font-size:22px;">Convite de Cadastro</h2>

                            <p style="font-size:16px; line-height:1.6;">Olá!</p>

                            <p style="font-size:16px; line-height:1.6;">
                                Você recebeu um convite para realizar seu cadastro no
                                <strong>Sistema de Lideranças</strong>.
                            </p>

                            <p style="font-size:16px; line-height:1.6;">
                                <strong>Liderança responsável:</strong> {nome_cabo}
                            </p>

                            <div style="text-align:center; margin:30px 0;">
                                <a href="{link}"
                                   style="background:#198754; color:#ffffff; padding:14px 28px; text-decoration:none; border-radius:6px; font-size:16px; display:inline-block;">
                                    Realizar Cadastro
                                </a>
                            </div>

                            <p style="font-size:14px; line-height:1.6; color:#4b5563;">
                                Caso o botão não funcione, copie e cole o link abaixo no navegador:
                            </p>

                            <p style="font-size:13px; line-height:1.6; word-break:break-all; color:#2563eb;">
                                {link}
                            </p>

                            <p style="font-size:14px; line-height:1.6; color:#6b7280;">
                                Se você não reconhece este convite, apenas ignore esta mensagem.
                            </p>
                        </td>
                    </tr>

                    <tr>
                        <td style="background:#f9fafb; padding:18px; text-align:center; color:#6b7280; font-size:12px;">
                            Esta é uma mensagem automática. Não responda este e-mail.
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>
</html>
"""

        msg = MIMEMultipart("alternative")
        msg["Subject"] = assunto
        msg["From"] = f"Sistema de Lideranças <{MAIL_FROM}>"
        msg["To"] = destinatario
        msg["Reply-To"] = MAIL_FROM

        msg.attach(MIMEText(corpo_texto, "plain", "utf-8"))
        msg.attach(MIMEText(corpo_html, "html", "utf-8"))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as servidor:
            servidor.starttls()
            servidor.login(SMTP_USER, SMTP_PASSWORD)
            servidor.sendmail(MAIL_FROM, [destinatario], msg.as_string())

        print(f"E-mail enviado com sucesso para {destinatario}")
        return True

    except Exception as erro:
        print("ERRO AO ENVIAR E-MAIL:", erro)
        return False

# ============================================================
# DASHBOARD / HOME
# ============================================================

@app.route("/")
@login_required
def home():
    """
    Rota principal do sistema.

    Responsável por carregar o painel inicial com:
    - total de lideranças;
    - total de apoiadores;
    - ranking de lideranças por quantidade de apoiadores;
    - filtro por Região Administrativa;
    - restrição de dados para usuário do tipo CABO.

    Regras:
    - ADMIN, DEPUTADO, CHEFE_GABINETE e SECRETARIA visualizam dados gerais.
    - CABO visualiza apenas os dados da própria liderança.
    """

    # Captura o filtro de região enviado pela tela.
    regiao = request.args.get("regiao", "").strip()

    # Recupera o tipo de acesso e o cabo vinculado ao usuário logado.
    tipo_acesso = session.get("tipo_acesso")
    cabo_sessao = session.get("cabo_id")

    # Valores padrão usados caso ocorra erro ou não haja dados.
    total_cabos = 0
    total_contatos = 0
    ranking_cabos = []
    regioes = []

    # Abre conexão com o Oracle.
    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return render_template(
            "dashboard.html",
            total_cabos=total_cabos,
            total_contatos=total_contatos,
            ranking_cabos=ranking_cabos,
            regiao=regiao,
            regioes=regioes
        )

    cursor = conexao.cursor()

    try:
        # ====================================================
        # 1. CARREGAR REGIÕES DISPONÍVEIS
        # ====================================================
        # Se o usuário for CABO, carrega apenas a região da própria liderança.
        # Caso contrário, carrega todas as regiões cadastradas.
        if tipo_acesso == "CABO":
            cursor.execute("""
                SELECT DISTINCT REGIAO_ADMINISTRATIVA
                FROM CABOS_ELEITORAIS
                WHERE ID = :1
                  AND REGIAO_ADMINISTRATIVA IS NOT NULL
                ORDER BY REGIAO_ADMINISTRATIVA
            """, (cabo_sessao,))
        else:
            cursor.execute("""
                SELECT DISTINCT REGIAO_ADMINISTRATIVA
                FROM CABOS_ELEITORAIS
                WHERE REGIAO_ADMINISTRATIVA IS NOT NULL
                ORDER BY REGIAO_ADMINISTRATIVA
            """)

        regioes = [r[0] for r in cursor.fetchall()]

        # ====================================================
        # 2. TOTAL DE LIDERANÇAS
        # ====================================================
        if tipo_acesso == "CABO":
            # CABO vê somente sua própria liderança.
            cursor.execute("""
                SELECT COUNT(*)
                FROM CABOS_ELEITORAIS
                WHERE ID = :1
            """, (cabo_sessao,))
        elif regiao:
            # Usuários administrativos podem filtrar por região.
            cursor.execute("""
                SELECT COUNT(*)
                FROM CABOS_ELEITORAIS
                WHERE REGIAO_ADMINISTRATIVA = :1
            """, (regiao,))
        else:
            # Usuários administrativos visualizam o total geral.
            cursor.execute("""
                SELECT COUNT(*)
                FROM CABOS_ELEITORAIS
            """)

        total_cabos = cursor.fetchone()[0]

        # ====================================================
        # 3. TOTAL DE APOIADORES
        # ====================================================
        if tipo_acesso == "CABO":
            # CABO vê apenas apoiadores vinculados a ele.
            cursor.execute("""
                SELECT COUNT(*)
                FROM CONTATOS_CAMPANHA
                WHERE CABO_ID = :1
            """, (cabo_sessao,))
        elif regiao:
            # Conta apoiadores vinculados a lideranças da região filtrada.
            cursor.execute("""
                SELECT COUNT(ct.ID)
                FROM CONTATOS_CAMPANHA ct
                JOIN CABOS_ELEITORAIS c ON c.ID = ct.CABO_ID
                WHERE c.REGIAO_ADMINISTRATIVA = :1
            """, (regiao,))
        else:
            # Total geral de apoiadores.
            cursor.execute("""
                SELECT COUNT(*)
                FROM CONTATOS_CAMPANHA
            """)

        total_contatos = cursor.fetchone()[0]

        # ====================================================
        # 4. RANKING DE LIDERANÇAS
        # ====================================================
        # Monta SQL dinâmico para respeitar filtros de região e perfil CABO.
        sql = """
            SELECT
                c.NOME,
                c.REGIAO_ADMINISTRATIVA,
                COUNT(ct.ID) AS TOTAL_CONTATOS
            FROM CABOS_ELEITORAIS c
            LEFT JOIN CONTATOS_CAMPANHA ct ON ct.CABO_ID = c.ID
            WHERE 1=1
        """

        params = {}

        if regiao:
            sql += " AND c.REGIAO_ADMINISTRATIVA = :regiao"
            params["regiao"] = regiao

        if tipo_acesso == "CABO":
            sql += " AND c.ID = :cabo_sessao"
            params["cabo_sessao"] = cabo_sessao

        sql += """
            GROUP BY c.NOME, c.REGIAO_ADMINISTRATIVA
            ORDER BY TOTAL_CONTATOS DESC, c.NOME
        """

        cursor.execute(sql, params)
        ranking_cabos = cursor.fetchall()

    except oracledb.Error as erro:
        # O erro técnico fica no log; o usuário recebe mensagem simples.
        print("ERRO DASHBOARD:", erro)
        flash("Erro ao carregar o dashboard.", "danger")

    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "dashboard.html",
        total_cabos=total_cabos,
        total_contatos=total_contatos,
        ranking_cabos=ranking_cabos,
        regiao=regiao,
        regioes=regioes
    )

# ============================================================
# LIDERANÇAS / CABOS
# ============================================================

@app.route("/cadastrar", methods=["GET", "POST"])
@login_required
@perfil_required("ADMIN", "CHEFE_GABINETE", "SECRETARIA")
def cadastrar():
    """
    Cadastra uma nova liderança no sistema.
    Apenas perfis administrativos podem acessar.
    """

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        email = request.form.get("email", "").strip().lower()
        telefone = request.form.get("telefone", "").strip()
        endereco = request.form.get("endereco", "").strip()
        regiao_administrativa = request.form.get("regiao_administrativa", "").strip()
        data_nascimento = converter_data(request.form.get("data_nascimento", "").strip())
        cep = request.form.get("cep", "").strip()

        if not nome:
            flash("O campo nome é obrigatório.", "warning")
            return render_template("cadastrar.html")

        conexao = conectar_oracle()
        if conexao is None:
            flash("Não foi possível conectar ao banco de dados.", "danger")
            return render_template("cadastrar.html")

        cursor = conexao.cursor()

        try:
            cursor.execute("""
                INSERT INTO CABOS_ELEITORAIS
                    (NOME, EMAIL, TELEFONE, ENDERECO, REGIAO_ADMINISTRATIVA, DATA_NASCIMENTO, CEP)
                VALUES
                    (:1, :2, :3, :4, :5, :6, :7)
            """, (
                nome,
                email,
                telefone,
                endereco,
                regiao_administrativa,
                data_nascimento,
                cep
            ))

            conexao.commit()
            flash("Liderança cadastrada com sucesso.", "success")
            return redirect(url_for("listar"))

        except oracledb.Error as erro:
            conexao.rollback()
            print("ERRO AO CADASTRAR LIDERANÇA:", erro)
            flash("Erro ao cadastrar liderança.", "danger")

        finally:
            cursor.close()
            conexao.close()

    return render_template("cadastrar.html")


@app.route("/listar")
@login_required
def listar():
    """
    Lista as lideranças cadastradas.
    Permite filtro por nome e região.
    Se o usuário for CABO, exibe apenas sua própria liderança.
    """

    termo = request.args.get("busca", "").strip()
    regiao = request.args.get("regiao", "").strip()

    tipo_acesso = session.get("tipo_acesso")
    cabo_sessao = session.get("cabo_id")

    registros = []
    regioes = []

    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return render_template(
            "listar.html",
            registros=registros,
            termo=termo,
            regiao=regiao,
            regioes=regioes
        )

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT DISTINCT REGIAO_ADMINISTRATIVA
            FROM CABOS_ELEITORAIS
            WHERE REGIAO_ADMINISTRATIVA IS NOT NULL
            ORDER BY REGIAO_ADMINISTRATIVA
        """)
        regioes = [r[0] for r in cursor.fetchall()]

        sql = """
            SELECT
                c.ID,
                c.NOME,
                c.EMAIL,
                c.TELEFONE,
                c.ENDERECO,
                c.REGIAO_ADMINISTRATIVA,
                c.DATA_NASCIMENTO,
                c.CEP,
                COUNT(ct.ID) AS TOTAL_CONTATOS,
                SUM(CASE WHEN ct.CONSENTIU_CONTATO = 'S' THEN 1 ELSE 0 END) AS TOTAL_SIM,
                SUM(CASE WHEN ct.CONSENTIU_CONTATO = 'N' THEN 1 ELSE 0 END) AS TOTAL_NAO
            FROM CABOS_ELEITORAIS c
            LEFT JOIN CONTATOS_CAMPANHA ct ON ct.CABO_ID = c.ID
            WHERE 1=1
        """

        params = {}

        if termo:
            sql += " AND TRIM(UPPER(c.NOME)) LIKE TRIM(UPPER(:busca))"
            params["busca"] = f"%{termo.upper()}%"

        if regiao:
            sql += " AND c.REGIAO_ADMINISTRATIVA = :regiao"
            params["regiao"] = regiao

        if tipo_acesso == "CABO":
            sql += " AND c.ID = :cabo_sessao"
            params["cabo_sessao"] = cabo_sessao

        sql += """
            GROUP BY
                c.ID,
                c.NOME,
                c.EMAIL,
                c.TELEFONE,
                c.ENDERECO,
                c.REGIAO_ADMINISTRATIVA,
                c.DATA_NASCIMENTO,
                c.CEP
            ORDER BY c.ID
        """

        cursor.execute(sql, params)
        registros = cursor.fetchall()

    except oracledb.Error as erro:
        print("ERRO NA LISTAGEM DE LIDERANÇAS:", erro)
        flash("Erro ao listar lideranças.", "danger")

    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "listar.html",
        registros=registros,
        termo=termo,
        regiao=regiao,
        regioes=regioes
    )


@app.route("/exportar-excel")
@login_required
def exportar_excel():
    """
    Exporta a lista de lideranças para Excel.
    Respeita os mesmos filtros da tela de listagem.
    """

    termo = request.args.get("busca", "").strip()
    regiao = request.args.get("regiao", "").strip()

    tipo_acesso = session.get("tipo_acesso")
    cabo_sessao = session.get("cabo_id")

    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()

    try:
        sql = """
            SELECT
                c.ID,
                c.NOME,
                c.EMAIL,
                c.TELEFONE,
                c.ENDERECO,
                c.REGIAO_ADMINISTRATIVA,
                c.DATA_NASCIMENTO,
                c.CEP,
                COUNT(ct.ID) AS TOTAL_CONTATOS
            FROM CABOS_ELEITORAIS c
            LEFT JOIN CONTATOS_CAMPANHA ct ON ct.CABO_ID = c.ID
            WHERE 1=1
        """

        params = {}

        if termo:
            sql += " AND TRIM(UPPER(c.NOME)) LIKE TRIM(UPPER(:busca))"
            params["busca"] = f"%{termo.upper()}%"

        if regiao:
            sql += " AND c.REGIAO_ADMINISTRATIVA = :regiao"
            params["regiao"] = regiao

        if tipo_acesso == "CABO":
            sql += " AND c.ID = :cabo_sessao"
            params["cabo_sessao"] = cabo_sessao

        sql += """
            GROUP BY
                c.ID,
                c.NOME,
                c.EMAIL,
                c.TELEFONE,
                c.ENDERECO,
                c.REGIAO_ADMINISTRATIVA,
                c.DATA_NASCIMENTO,
                c.CEP
            ORDER BY c.ID
        """

        cursor.execute(sql, params)
        registros = cursor.fetchall()

    except oracledb.Error as erro:
        print("ERRO AO EXPORTAR LIDERANÇAS:", erro)
        flash("Erro ao exportar lideranças.", "danger")
        return redirect(url_for("listar"))

    finally:
        cursor.close()
        conexao.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lideranças"

    sheet.append([
        "ID",
        "Nome",
        "Email",
        "Telefone",
        "Endereço",
        "Região Administrativa",
        "Data de Nascimento",
        "CEP",
        "Total de Contatos"
    ])

    for registro in registros:
        linha = list(registro)

        if linha[6]:
            linha[6] = linha[6].strftime("%d/%m/%Y")

        sheet.append(linha)

    arquivo_excel = BytesIO()
    workbook.save(arquivo_excel)
    arquivo_excel.seek(0)

    return send_file(
        arquivo_excel,
        as_attachment=True,
        download_name="liderancas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/editar/<int:id>", methods=["GET", "POST"])
@login_required
@perfil_required("ADMIN", "CHEFE_GABINETE", "SECRETARIA")
def editar(id):
    """
    Edita os dados de uma liderança.
    """

    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()

    try:
        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            email = request.form.get("email", "").strip().lower()
            telefone = request.form.get("telefone", "").strip()
            endereco = request.form.get("endereco", "").strip()
            regiao_administrativa = request.form.get("regiao_administrativa", "").strip()
            data_nascimento = converter_data(request.form.get("data_nascimento", "").strip())
            cep = request.form.get("cep", "").strip()

            cursor.execute("""
                UPDATE CABOS_ELEITORAIS
                SET NOME = :1,
                    EMAIL = :2,
                    TELEFONE = :3,
                    ENDERECO = :4,
                    REGIAO_ADMINISTRATIVA = :5,
                    DATA_NASCIMENTO = :6,
                    CEP = :7
                WHERE ID = :8
            """, (
                nome,
                email,
                telefone,
                endereco,
                regiao_administrativa,
                data_nascimento,
                cep,
                id
            ))

            conexao.commit()
            flash("Liderança atualizada com sucesso.", "success")
            return redirect(url_for("listar"))

        cursor.execute("""
            SELECT
                ID,
                NOME,
                EMAIL,
                TELEFONE,
                ENDERECO,
                REGIAO_ADMINISTRATIVA,
                DATA_NASCIMENTO,
                CEP
            FROM CABOS_ELEITORAIS
            WHERE ID = :1
        """, (id,))

        registro = cursor.fetchone()

        if not registro:
            flash("Liderança não encontrada.", "warning")
            return redirect(url_for("listar"))

    except oracledb.Error as erro:
        conexao.rollback()
        print("ERRO AO EDITAR LIDERANÇA:", erro)
        flash("Erro ao editar liderança.", "danger")
        return redirect(url_for("listar"))

    finally:
        cursor.close()
        conexao.close()

    return render_template("editar.html", registro=registro)


@app.route("/excluir/<int:id>")
@login_required
@perfil_required("ADMIN", "CHEFE_GABINETE")
def excluir(id):
    """
    Exclui uma liderança somente se não houver apoiadores ou convites ativos vinculados.
    Convites com status CANCELADO são removidos antes da validação.
    """

    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT COUNT(*)
            FROM CONTATOS_CAMPANHA
            WHERE CABO_ID = :1
        """, (id,))
        total_contatos = cursor.fetchone()[0]

        cursor.execute("""
            DELETE FROM CONVITES_CONTATO
            WHERE CABO_ID = :1
              AND STATUS = 'CANCELADO'
        """, (id,))

        cursor.execute("""
            SELECT COUNT(*)
            FROM CONVITES_CONTATO
            WHERE CABO_ID = :1
        """, (id,))
        total_convites = cursor.fetchone()[0]

        if total_contatos > 0 or total_convites > 0:
            conexao.rollback()
            flash(
                f"Não é possível excluir esta liderança. "
                f"Existem {total_contatos} apoiador(es) e "
                f"{total_convites} convite(s) ativo(s)/histórico(s) vinculados a ela.",
                "warning"
            )
            return redirect(url_for("listar"))

        cursor.execute("""
            DELETE FROM CABOS_ELEITORAIS
            WHERE ID = :1
        """, (id,))

        conexao.commit()
        flash("Liderança excluída com sucesso.", "success")

    except oracledb.Error as erro:
        conexao.rollback()
        print("ERRO AO EXCLUIR LIDERANÇA:", erro)
        flash("Erro ao excluir liderança.", "danger")

    finally:
        cursor.close()
        conexao.close()

    return redirect(url_for("listar"))

# ============================================================
# CONTATOS / APOIADORES
# ============================================================

@app.route("/cadastrar-contato", methods=["GET", "POST"])
@login_required
@perfil_required("ADMIN", "CHEFE_GABINETE", "SECRETARIA")
def cadastrar_contato():
    """
    Cadastra manualmente um apoiador no sistema.

    Regras:
    - Apenas perfis administrativos podem cadastrar apoiadores manualmente.
    - Todo apoiador precisa estar vinculado a uma liderança.
    - Nome e liderança são obrigatórios.
    """

    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("home"))

    cursor = conexao.cursor()

    try:
        # Carrega as lideranças para preencher o select do formulário.
        cursor.execute("""
            SELECT ID, NOME
            FROM CABOS_ELEITORAIS
            ORDER BY NOME
        """)
        cabos = cursor.fetchall()

        # Se o formulário foi enviado, inicia o cadastro.
        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            email = request.form.get("email", "").strip().lower()
            telefone = request.form.get("telefone", "").strip()
            endereco = request.form.get("endereco", "").strip()
            regiao = request.form.get("regiao_administrativa", "").strip()
            cabo_id = request.form.get("cabo_id", "").strip()
            consentiu = request.form.get("consentiu_contato", "N")
            observacao = request.form.get("observacao", "").strip()
            data_nascimento = converter_data(request.form.get("data_nascimento", "").strip())
            cep = request.form.get("cep", "").strip()

            # Validação mínima obrigatória.
            if not nome or not cabo_id:
                flash("Nome e liderança responsável são obrigatórios.", "warning")
                return render_template("cadastrar_contato.html", cabos=cabos)

            # Evita duplicidade de e-mail, caso informado.
            if email:
                cursor.execute("""
                    SELECT ID
                    FROM CONTATOS_CAMPANHA
                    WHERE LOWER(EMAIL) = LOWER(:1)
                """, (email,))
                contato_existente = cursor.fetchone()

                if contato_existente:
                    flash("Já existe um apoiador cadastrado com esse e-mail.", "warning")
                    return render_template("cadastrar_contato.html", cabos=cabos)

            # Insere o apoiador vinculado à liderança selecionada.
            cursor.execute("""
                INSERT INTO CONTATOS_CAMPANHA
                    (
                        NOME,
                        EMAIL,
                        TELEFONE,
                        ENDERECO,
                        REGIAO_ADMINISTRATIVA,
                        CABO_ID,
                        CONSENTIU_CONTATO,
                        OBSERVACAO,
                        DATA_NASCIMENTO,
                        CEP
                    )
                VALUES
                    (:1, :2, :3, :4, :5, :6, :7, :8, :9, :10)
            """, (
                nome,
                email,
                telefone,
                endereco,
                regiao,
                int(cabo_id),
                consentiu,
                observacao,
                data_nascimento,
                cep
            ))

            conexao.commit()

            flash("Apoiador cadastrado com sucesso.", "success")
            return redirect(url_for("listar_contatos"))

    except oracledb.Error as erro:
        conexao.rollback()
        print("ERRO AO CADASTRAR APOIADOR:", erro)
        flash("Erro ao cadastrar apoiador.", "danger")

    finally:
        cursor.close()
        conexao.close()

    return render_template("cadastrar_contato.html", cabos=cabos)

# ============================================================
# LISTAGEM DE APOIADORES
# ============================================================

@app.route("/listar-contatos")
@login_required
def listar_contatos():
    """
    Lista os apoiadores cadastrados no sistema.

    Funcionalidades:
    - filtro por nome;
    - filtro por liderança;
    - filtro por região administrativa;
    - restrição automática para usuários do tipo CABO.
    """

    # Captura os filtros enviados pela URL.
    termo = request.args.get("busca", "").strip()
    cabo_id = request.args.get("cabo_id", "").strip()
    regiao = request.args.get("regiao", "").strip()

    # Dados do usuário logado.
    tipo_acesso = session.get("tipo_acesso")
    cabo_sessao = session.get("cabo_id")

    registros = []
    cabos = []
    regioes = []

    # Conecta ao Oracle.
    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")

        return render_template(
            "listar_contatos.html",
            registros=registros,
            termo=termo,
            cabo_id=cabo_id,
            regiao=regiao,
            cabos=cabos,
            regioes=regioes
        )

    cursor = conexao.cursor()

    try:

        # ====================================================
        # CARREGA AS LIDERANÇAS PARA O FILTRO
        # ====================================================
        # Usuário CABO visualiza apenas sua própria liderança.
        if tipo_acesso == "CABO":

            cursor.execute("""
                SELECT ID, NOME
                FROM CABOS_ELEITORAIS
                WHERE ID = :1
                ORDER BY NOME
            """, (cabo_sessao,))

        else:

            cursor.execute("""
                SELECT ID, NOME
                FROM CABOS_ELEITORAIS
                ORDER BY NOME
            """)

        cabos = cursor.fetchall()

        # ====================================================
        # CARREGA AS REGIÕES DISPONÍVEIS
        # ====================================================
        cursor.execute("""
            SELECT DISTINCT REGIAO_ADMINISTRATIVA
            FROM CONTATOS_CAMPANHA
            WHERE REGIAO_ADMINISTRATIVA IS NOT NULL
            ORDER BY REGIAO_ADMINISTRATIVA
        """)

        regioes = [r[0] for r in cursor.fetchall()]

        # ====================================================
        # SQL BASE DA LISTAGEM
        # ====================================================
        sql = """
            SELECT
                ct.ID,
                ct.NOME,
                ct.EMAIL,
                ct.TELEFONE,
                ct.REGIAO_ADMINISTRATIVA,
                c.NOME AS CABO_NOME,
                ct.DATA_NASCIMENTO,
                ct.CEP,
                ct.CONSENTIU_CONTATO
            FROM CONTATOS_CAMPANHA ct
            JOIN CABOS_ELEITORAIS c
                ON c.ID = ct.CABO_ID
            WHERE 1=1
        """

        params = {}

        # ====================================================
        # FILTRO POR NOME
        # ====================================================
        if termo:
            sql += """
                AND TRIM(UPPER(ct.NOME))
                    LIKE TRIM(UPPER(:busca))
            """
            params["busca"] = f"%{termo}%"

        # ====================================================
        # FILTRO POR LIDERANÇA
        # ====================================================
        if cabo_id:
            sql += " AND ct.CABO_ID = :cabo_id"
            params["cabo_id"] = int(cabo_id)

        # ====================================================
        # FILTRO POR REGIÃO
        # ====================================================
        if regiao:
            sql += """
                AND ct.REGIAO_ADMINISTRATIVA = :regiao
            """
            params["regiao"] = regiao

        # ====================================================
        # RESTRIÇÃO PARA PERFIL CABO
        # ====================================================
        # Usuário CABO só visualiza seus próprios apoiadores.
        if tipo_acesso == "CABO":
            sql += " AND ct.CABO_ID = :cabo_sessao"
            params["cabo_sessao"] = cabo_sessao

        # ====================================================
        # ORDENAÇÃO FINAL
        # ====================================================
        sql += """
            ORDER BY
                ct.NOME
        """

        cursor.execute(sql, params)

        registros = cursor.fetchall()

    except oracledb.Error as erro:

        print("ERRO LISTAR APOIADORES:", erro)

        flash(
            "Erro ao carregar os apoiadores.",
            "danger"
        )

    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "listar_contatos.html",
        registros=registros,
        termo=termo,
        cabo_id=cabo_id,
        regiao=regiao,
        cabos=cabos,
        regioes=regioes
    )

# ============================================================
# EXPORTAÇÃO DE APOIADORES PARA EXCEL
# ============================================================

@app.route("/exportar-contatos-excel")
@login_required
def exportar_contatos_excel():
    """
    Exporta os apoiadores para um arquivo Excel.

    Funcionalidades:
    - respeita os filtros da tela de listagem;
    - exporta por nome, liderança e região;
    - restringe automaticamente os dados para usuário do tipo CABO.
    """

    termo = request.args.get("busca", "").strip()
    cabo_id = request.args.get("cabo_id", "").strip()
    regiao = request.args.get("regiao", "").strip()

    tipo_acesso = session.get("tipo_acesso")
    cabo_sessao = session.get("cabo_id")

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar_contatos"))

    cursor = conexao.cursor()

    try:
        sql = """
            SELECT
                ct.ID,
                ct.NOME,
                ct.EMAIL,
                ct.TELEFONE,
                ct.ENDERECO,
                ct.REGIAO_ADMINISTRATIVA,
                c.NOME AS CABO_NOME,
                ct.CONSENTIU_CONTATO,
                ct.OBSERVACAO,
                ct.DATA_CADASTRO,
                ct.DATA_NASCIMENTO,
                ct.CEP
            FROM CONTATOS_CAMPANHA ct
            JOIN CABOS_ELEITORAIS c
                ON c.ID = ct.CABO_ID
            WHERE 1=1
        """

        params = {}

        if termo:
            sql += """
                AND TRIM(UPPER(ct.NOME))
                    LIKE TRIM(UPPER(:busca))
            """
            params["busca"] = f"%{termo}%"

        if cabo_id:
            sql += " AND ct.CABO_ID = :cabo_id"
            params["cabo_id"] = int(cabo_id)

        if regiao:
            sql += " AND ct.REGIAO_ADMINISTRATIVA = :regiao"
            params["regiao"] = regiao

        if tipo_acesso == "CABO":
            sql += " AND ct.CABO_ID = :cabo_sessao"
            params["cabo_sessao"] = cabo_sessao

        sql += """
            ORDER BY ct.NOME
        """

        cursor.execute(sql, params)
        registros = cursor.fetchall()

    except oracledb.Error as erro:
        print("ERRO AO EXPORTAR APOIADORES:", erro)
        flash("Erro ao exportar apoiadores.", "danger")
        return redirect(url_for("listar_contatos"))

    finally:
        cursor.close()
        conexao.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Apoiadores"

    sheet.append([
        "ID",
        "Nome",
        "Email",
        "Telefone",
        "Endereço",
        "Região Administrativa",
        "Liderança",
        "Consentiu Contato",
        "Observação",
        "Data Cadastro",
        "Data Nascimento",
        "CEP"
    ])

    for registro in registros:
        linha = list(registro)

        if linha[9]:
            linha[9] = linha[9].strftime("%d/%m/%Y")

        if linha[10]:
            linha[10] = linha[10].strftime("%d/%m/%Y")

        sheet.append(linha)

    arquivo_excel = BytesIO()
    workbook.save(arquivo_excel)
    arquivo_excel.seek(0)

    return send_file(
        arquivo_excel,
        as_attachment=True,
        download_name="apoiadores.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ============================================================
# EDITAR APOIADOR
# ============================================================

@app.route("/editar-contato/<int:id>", methods=["GET", "POST"])
@login_required
def editar_contato(id):
    """
    Edita os dados de um apoiador já cadastrado.

    Funcionalidades:
    - carrega as lideranças para o select;
    - busca os dados atuais do apoiador;
    - valida nome e liderança;
    - impede duplicidade de e-mail em outro apoiador;
    - atualiza os dados no banco.
    """

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar_contatos"))

    cursor = conexao.cursor()

    try:
        # Carrega as lideranças para o campo de seleção.
        cursor.execute("""
            SELECT ID, NOME
            FROM CABOS_ELEITORAIS
            ORDER BY NOME
        """)
        cabos = cursor.fetchall()

        # Busca o apoiador que será editado.
        cursor.execute("""
            SELECT
                ID,
                NOME,
                EMAIL,
                TELEFONE,
                ENDERECO,
                REGIAO_ADMINISTRATIVA,
                CABO_ID,
                CONSENTIU_CONTATO,
                OBSERVACAO,
                DATA_NASCIMENTO,
                CEP
            FROM CONTATOS_CAMPANHA
            WHERE ID = :1
        """, (id,))
        contato = cursor.fetchone()

        if not contato:
            flash("Apoiador não encontrado.", "warning")
            return redirect(url_for("listar_contatos"))

        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            email = request.form.get("email", "").strip().lower()
            telefone = request.form.get("telefone", "").strip()
            endereco = request.form.get("endereco", "").strip()
            regiao = request.form.get("regiao_administrativa", "").strip()
            cabo_id = request.form.get("cabo_id", "").strip()
            consentiu = request.form.get("consentiu_contato", "N")
            observacao = request.form.get("observacao", "").strip()
            data_nascimento = converter_data(
                request.form.get("data_nascimento", "").strip()
            )
            cep = request.form.get("cep", "").strip()

            # Validação obrigatória.
            if not nome or not cabo_id:
                flash("Nome e liderança responsável são obrigatórios.", "warning")
                return render_template(
                    "editar_contato.html",
                    contato=contato,
                    cabos=cabos
                )

            # Verifica se o e-mail já pertence a outro apoiador.
            if email:
                cursor.execute("""
                    SELECT ID
                    FROM CONTATOS_CAMPANHA
                    WHERE LOWER(EMAIL) = LOWER(:1)
                      AND ID <> :2
                """, (email, id))

                contato_email_existente = cursor.fetchone()

                if contato_email_existente:
                    flash(
                        "Já existe outro apoiador cadastrado com esse e-mail.",
                        "warning"
                    )
                    return render_template(
                        "editar_contato.html",
                        contato=contato,
                        cabos=cabos
                    )

            # Atualiza o apoiador no banco.
            cursor.execute("""
                UPDATE CONTATOS_CAMPANHA
                SET NOME = :1,
                    EMAIL = :2,
                    TELEFONE = :3,
                    ENDERECO = :4,
                    REGIAO_ADMINISTRATIVA = :5,
                    CABO_ID = :6,
                    CONSENTIU_CONTATO = :7,
                    OBSERVACAO = :8,
                    DATA_NASCIMENTO = :9,
                    CEP = :10
                WHERE ID = :11
            """, (
                nome,
                email,
                telefone,
                endereco,
                regiao,
                int(cabo_id),
                consentiu,
                observacao,
                data_nascimento,
                cep,
                id
            ))

            conexao.commit()

            flash("Apoiador atualizado com sucesso.", "success")
            return redirect(url_for("listar_contatos"))

    except oracledb.Error as erro:
        conexao.rollback()
        print("ERRO AO EDITAR APOIADOR:", erro)
        flash("Erro ao carregar ou atualizar apoiador.", "danger")
        return redirect(url_for("listar_contatos"))

    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "editar_contato.html",
        contato=contato,
        cabos=cabos
    )

# ============================================================
# EXCLUIR APOIADOR
# ============================================================

@app.route("/excluir-contato/<int:id>")
@login_required
@perfil_required("ADMIN", "CHEFE_GABINETE", "SECRETARIA")
def excluir_contato(id):
    """
    Exclui um apoiador do sistema.

    Apenas usuários administrativos podem excluir.
    """

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar_contatos"))

    cursor = conexao.cursor()

    try:
        # Verifica se o apoiador existe.
        cursor.execute("""
            SELECT ID, NOME
            FROM CONTATOS_CAMPANHA
            WHERE ID = :1
        """, (id,))

        contato = cursor.fetchone()

        if not contato:
            flash("Apoiador não encontrado.", "warning")
            return redirect(url_for("listar_contatos"))

        # Exclui o apoiador.
        cursor.execute("""
            DELETE FROM CONTATOS_CAMPANHA
            WHERE ID = :1
        """, (id,))

        conexao.commit()
        
        registrar_auditoria(
        acao="EXCLUIR_CONTATO",
        tabela_afetada="CONTATOS_CAMPANHA",
        registro_id=id,
        descricao=f"Contato ID {id} excluído do sistema.")

        flash("Apoiador excluído com sucesso.", "success")

    except oracledb.Error as erro:
        conexao.rollback()
        print("ERRO AO EXCLUIR APOIADOR:", erro)

        flash(
            "Erro ao excluir apoiador.",
            "danger"
        )

    finally:
        cursor.close()
        conexao.close()

    return redirect(url_for("listar_contatos"))


# ============================================================
# CONTATOS / APOIADORES DA LIDERANÇA
# ============================================================

@app.route("/cabo/<int:cabo_id>/contatos")
@login_required
def contatos_do_cabo(cabo_id):
    """
    Exibe os apoiadores vinculados a uma liderança específica.

    Regras:
    - usuário CABO só pode visualizar seus próprios apoiadores;
    - usuários administrativos visualizam qualquer liderança.
    """

    tipo_acesso = session.get("tipo_acesso")
    cabo_sessao = session.get("cabo_id")

    # Segurança:
    # Usuário CABO não pode acessar contatos de outra liderança.
    if tipo_acesso == "CABO" and cabo_sessao != cabo_id:
        flash(
            "Você só pode acessar os seus próprios apoiadores.",
            "danger"
        )
        return redirect(url_for("home"))

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()

    try:
        # Busca os dados da liderança.
        cursor.execute("""
            SELECT
                ID,
                NOME,
                REGIAO_ADMINISTRATIVA
            FROM CABOS_ELEITORAIS
            WHERE ID = :1
        """, (cabo_id,))

        cabo = cursor.fetchone()

        if not cabo:
            flash("Liderança não encontrada.", "warning")
            return redirect(url_for("listar"))

        # Busca os apoiadores vinculados à liderança.
        cursor.execute("""
            SELECT
                ID,                    -- item[0]
                NOME,                  -- item[1]
                TELEFONE,              -- item[2]
                ENDERECO,              -- item[3]
                DATA_NASCIMENTO,       -- item[4]
                CEP,                   -- item[5]
                REGIAO_ADMINISTRATIVA, -- item[6]
                CONSENTIU_CONTATO,     -- item[7]
                OBSERVACAO,            -- item[8]
                DATA_CADASTRO          -- item[9]
            FROM CONTATOS_CAMPANHA
            WHERE CABO_ID = :1
            ORDER BY NOME
        """, (cabo_id,))

        contatos = cursor.fetchall()

        # Total de apoiadores vinculados.
        total_contatos = len(contatos)

    except oracledb.Error as erro:

        print("ERRO CONTATOS DA LIDERANÇA:", erro)

        flash(
            "Erro ao carregar os apoiadores da liderança.",
            "danger"
        )

        return redirect(url_for("listar"))

    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "contatos_do_cabo.html",
        cabo=cabo,
        contatos=contatos,
        total_contatos=total_contatos
    )

# ============================================================
# EXPORTAR APOIADORES DE UMA LIDERANÇA PARA EXCEL
# ============================================================

@app.route("/cabo/<int:cabo_id>/exportar-excel")
@login_required
def exportar_contatos_do_cabo(cabo_id):
    """
    Exporta para Excel todos os apoiadores vinculados a uma liderança específica.

    Regras:
    - ADMIN, CHEFE_GABINETE, SECRETARIA e DEPUTADO podem exportar qualquer liderança.
    - CABO só pode exportar os próprios apoiadores.
    """

    tipo_acesso = session.get("tipo_acesso")
    cabo_sessao = session.get("cabo_id")

    if tipo_acesso == "CABO" and cabo_sessao != cabo_id:
        flash("Você só pode exportar os seus próprios apoiadores.", "danger")
        return redirect(url_for("home"))

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT NOME
            FROM CABOS_ELEITORAIS
            WHERE ID = :1
        """, (cabo_id,))

        cabo = cursor.fetchone()

        if not cabo:
            flash("Liderança não encontrada.", "warning")
            return redirect(url_for("listar"))

        nome_cabo = cabo[0]

        cursor.execute("""
            SELECT
                ID,
                NOME,
                TELEFONE,
                ENDERECO,
                REGIAO_ADMINISTRATIVA,
                CONSENTIU_CONTATO,
                OBSERVACAO,
                DATA_CADASTRO,
                DATA_NASCIMENTO,
                CEP
            FROM CONTATOS_CAMPANHA
            WHERE CABO_ID = :1
            ORDER BY NOME
        """, (cabo_id,))

        contatos = cursor.fetchall()

    except oracledb.Error as erro:
        print("ERRO AO EXPORTAR APOIADORES DA LIDERANÇA:", erro)
        flash("Erro ao exportar apoiadores da liderança.", "danger")
        return redirect(url_for("listar"))

    finally:
        cursor.close()
        conexao.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Apoiadores"

    sheet.append([
        "ID",
        "Nome",
        "Telefone",
        "Endereço",
        "Região Administrativa",
        "Consentiu Contato",
        "Observação",
        "Data Cadastro",
        "Data Nascimento",
        "CEP"
    ])

    for contato in contatos:
        linha = list(contato)

        if linha[7]:
            linha[7] = linha[7].strftime("%d/%m/%Y")

        if linha[8]:
            linha[8] = linha[8].strftime("%d/%m/%Y")

        sheet.append(linha)

    arquivo_excel = BytesIO()
    workbook.save(arquivo_excel)
    arquivo_excel.seek(0)

    nome_seguro = nome_cabo.replace(" ", "_").lower()
    nome_arquivo = f"apoiadores_{nome_seguro}.xlsx"

    return send_file(
        arquivo_excel,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ============================================================
# RELATÓRIO GERAL
# ============================================================

@app.route("/relatorio-geral")
@login_required
@perfil_required("ADMIN", "DEPUTADO", "CHEFE_GABINETE")
def relatorio_geral():
    """
    Exibe o relatório geral do sistema.

    Funcionalidades:
    - total de lideranças cadastradas;
    - total de apoiadores cadastrados;
    - total de apoiadores por região;
    - ranking de lideranças por quantidade de apoiadores.

    Acesso permitido:
    - ADMIN
    - DEPUTADO
    - CHEFE_GABINETE
    """

    total_cabos = 0
    total_contatos = 0
    contatos_por_regiao = []
    ranking_cabos = []

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("home"))

    cursor = conexao.cursor()

    try:
        # Total geral de lideranças.
        cursor.execute("""
            SELECT COUNT(*)
            FROM CABOS_ELEITORAIS
        """)
        total_cabos = cursor.fetchone()[0]

        # Total geral de apoiadores.
        cursor.execute("""
            SELECT COUNT(*)
            FROM CONTATOS_CAMPANHA
        """)
        total_contatos = cursor.fetchone()[0]

        # Total de apoiadores agrupados por região administrativa.
        cursor.execute("""
            SELECT
                NVL(REGIAO_ADMINISTRATIVA, 'Não informada') AS REGIAO,
                COUNT(*) AS TOTAL
            FROM CONTATOS_CAMPANHA
            GROUP BY NVL(REGIAO_ADMINISTRATIVA, 'Não informada')
            ORDER BY TOTAL DESC, REGIAO
        """)
        contatos_por_regiao = cursor.fetchall()

        # Ranking de lideranças por quantidade de apoiadores.
        cursor.execute("""
            SELECT
                c.NOME,
                NVL(c.REGIAO_ADMINISTRATIVA, 'Não informada') AS REGIAO,
                COUNT(ct.ID) AS TOTAL_CONTATOS
            FROM CABOS_ELEITORAIS c
            LEFT JOIN CONTATOS_CAMPANHA ct
                ON ct.CABO_ID = c.ID
            GROUP BY
                c.NOME,
                NVL(c.REGIAO_ADMINISTRATIVA, 'Não informada')
            ORDER BY
                TOTAL_CONTATOS DESC,
                c.NOME
        """)
        ranking_cabos = cursor.fetchall()

    except oracledb.Error as erro:
        print("ERRO RELATÓRIO GERAL:", erro)
        flash("Erro ao carregar relatório geral.", "danger")
        return redirect(url_for("home"))

    finally:
        cursor.close()
        conexao.close()

    # Dados preparados para gráficos no template.
    regioes_labels = [item[0] for item in contatos_por_regiao]
    regioes_valores = [item[1] for item in contatos_por_regiao]

    cabos_labels = [item[0] for item in ranking_cabos]
    cabos_valores = [item[2] for item in ranking_cabos]

    return render_template(
        "relatorio_geral.html",
        total_cabos=total_cabos,
        total_contatos=total_contatos,
        contatos_por_regiao=contatos_por_regiao,
        ranking_cabos=ranking_cabos,
        regioes_labels=regioes_labels,
        regioes_valores=regioes_valores,
        cabos_labels=cabos_labels,
        cabos_valores=cabos_valores
    )

# ============================================================
# EXPORTAR RELATÓRIO GERAL PARA EXCEL
# ============================================================

@app.route("/exportar-relatorio-geral-excel")
@login_required
@perfil_required("ADMIN", "DEPUTADO", "CHEFE_GABINETE")
def exportar_relatorio_geral_excel():
    """
    Exporta o relatório geral do sistema para Excel.

    O arquivo terá duas abas:
    1. Ranking de Lideranças;
    2. Apoiadores por Região.

    Acesso permitido:
    - ADMIN
    - DEPUTADO
    - CHEFE_GABINETE
    """

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("relatorio_geral"))

    cursor = conexao.cursor()

    try:
        # Aba 1: ranking de lideranças por quantidade de apoiadores.
        cursor.execute("""
            SELECT
                c.NOME,
                NVL(c.REGIAO_ADMINISTRATIVA, 'Não informada') AS REGIAO,
                COUNT(ct.ID) AS TOTAL_CONTATOS
            FROM CABOS_ELEITORAIS c
            LEFT JOIN CONTATOS_CAMPANHA ct
                ON ct.CABO_ID = c.ID
            GROUP BY
                c.NOME,
                NVL(c.REGIAO_ADMINISTRATIVA, 'Não informada')
            ORDER BY
                TOTAL_CONTATOS DESC,
                c.NOME
        """)
        ranking_cabos = cursor.fetchall()

        # Aba 2: total de apoiadores por região administrativa.
        cursor.execute("""
            SELECT
                NVL(REGIAO_ADMINISTRATIVA, 'Não informada') AS REGIAO,
                COUNT(*) AS TOTAL
            FROM CONTATOS_CAMPANHA
            GROUP BY
                NVL(REGIAO_ADMINISTRATIVA, 'Não informada')
            ORDER BY
                TOTAL DESC,
                REGIAO
        """)
        contatos_por_regiao = cursor.fetchall()

    except oracledb.Error as erro:
        print("ERRO AO EXPORTAR RELATÓRIO GERAL:", erro)
        flash("Erro ao exportar relatório geral.", "danger")
        return redirect(url_for("relatorio_geral"))

    finally:
        cursor.close()
        conexao.close()

    workbook = Workbook()

    # Primeira aba: ranking de lideranças.
    aba1 = workbook.active
    aba1.title = "Ranking Liderancas"

    aba1.append([
        "Liderança",
        "Região",
        "Total de Apoiadores"
    ])

    for linha in ranking_cabos:
        aba1.append(list(linha))

    # Segunda aba: apoiadores por região.
    aba2 = workbook.create_sheet(title="Apoiadores por Regiao")

    aba2.append([
        "Região",
        "Total de Apoiadores"
    ])

    for linha in contatos_por_regiao:
        aba2.append(list(linha))

    arquivo_excel = BytesIO()
    workbook.save(arquivo_excel)
    arquivo_excel.seek(0)

    return send_file(
        arquivo_excel,
        as_attachment=True,
        download_name="relatorio_geral.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ============================================================
# LOGIN / LOGOUT
# ============================================================

@app.route("/login", methods=["GET", "POST"])
def login():
    """
    Realiza a autenticação do usuário.

    Fluxo:
    - recebe e-mail e senha;
    - consulta o usuário na tabela USUARIOS_SISTEMA;
    - verifica se está ativo;
    - valida a senha usando check_password_hash;
    - grava os dados principais na sessão Flask.
    """

    if request.method == "POST":
        email = request.form.get("usuario", "").strip().lower()
        senha = request.form.get("senha", "").strip()

        if not email or not senha:
            flash("Informe usuário e senha.", "warning")
            return render_template("login.html")

        conexao = conectar_oracle()

        if conexao is None:
            flash("Não foi possível conectar ao banco de dados.", "danger")
            return render_template("login.html")

        cursor = conexao.cursor()

        try:
            cursor.execute("""
                SELECT
                    ID,
                    NOME,
                    EMAIL,
                    SENHA_HASH,
                    TIPO_ACESSO,
                    CABO_ID,
                    ATIVO
                FROM USUARIOS_SISTEMA
                WHERE LOWER(EMAIL) = LOWER(:1)
            """, (email,))

            usuario = cursor.fetchone()

            if not usuario:
                flash("Usuário ou senha inválidos.", "danger")
                return render_template("login.html")

            usuario_id = usuario[0]
            nome = usuario[1]
            email_db = usuario[2]
            senha_hash = usuario[3]
            tipo_acesso = usuario[4]
            cabo_id = usuario[5]
            ativo = usuario[6]

            if ativo != "S":
                flash("Usuário inativo. Procure o administrador.", "warning")
                return render_template("login.html")

            if not senha_hash or not check_password_hash(senha_hash, senha):
                flash("Usuário ou senha inválidos.", "danger")
                return render_template("login.html")
            
            # =========================
            # LOGIN REALIZADO COM SUCESSO
            # =========================

            session.clear()
            
            session.permanent = True
            
            session["usuario_logado"] = True
            session["usuario_id"] = usuario_id
            session["usuario_nome"] = nome
            session["usuario_email"] = email_db
            session["tipo_acesso"] = tipo_acesso
            session["cabo_id"] = cabo_id

            flash("Login realizado com sucesso.", "success")
            return redirect(url_for("home"))

        except oracledb.Error as erro:
            print("ERRO LOGIN:", erro)
            flash("Erro ao realizar login.", "danger")

        finally:
            cursor.close()
            conexao.close()
            
        registrar_auditoria(
        acao="LOGIN",
        tabela_afetada="USUARIOS_SISTEMA",
        registro_id=usuario_id,
        descricao=f"Usuário {nome} realizou login no sistema."
    )
    return render_template("login.html")

@app.route("/logout")
def logout():

    registrar_auditoria(
        acao="LOGOUT",
        tabela_afetada="USUARIOS_SISTEMA",
        registro_id=session.get("usuario_id"),
        descricao=f"Usuário {session.get('usuario_nome')} saiu do sistema."
    )

    session.clear()

    flash("Logout realizado com sucesso.", "success")

    return redirect(url_for("login"))

# ============================================================
# GERENCIAMENTO DE USUÁRIOS
# ============================================================

@app.route("/cadastrar-usuario", methods=["GET", "POST"])
@login_required
@perfil_required("ADMIN")
def cadastrar_usuario():
    """
    Cadastra novos usuários do sistema.

    Regras:
    - apenas ADMIN pode cadastrar usuários;
    - e-mail não pode ser duplicado;
    - senha é salva com hash;
    - usuário do tipo CABO precisa estar vinculado a uma liderança.
    """

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("home"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT ID, NOME
            FROM CABOS_ELEITORAIS
            ORDER BY NOME
        """)
        cabos = cursor.fetchall()

        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            email = request.form.get("email", "").strip().lower()
            senha = request.form.get("senha", "").strip()
            tipo_acesso = request.form.get("tipo_acesso", "").strip()
            cabo_id = request.form.get("cabo_id", "").strip()
            ativo = request.form.get("ativo", "S").strip()

            if not nome or not email or not senha or not tipo_acesso:
                flash("Preencha os campos obrigatórios.", "warning")
                return render_template("cadastrar_usuario.html", cabos=cabos)

            cursor.execute("""
                SELECT ID
                FROM USUARIOS_SISTEMA
                WHERE LOWER(EMAIL) = LOWER(:1)
            """, (email,))
            usuario_existente = cursor.fetchone()

            if usuario_existente:
                flash("Já existe um usuário cadastrado com esse e-mail.", "warning")
                return render_template("cadastrar_usuario.html", cabos=cabos)

            if tipo_acesso == "CABO" and not cabo_id:
                flash("Para usuário do tipo CABO, selecione a liderança.", "warning")
                return render_template("cadastrar_usuario.html", cabos=cabos)

            senha_hash = generate_password_hash(senha)

            cursor.execute("""
                INSERT INTO USUARIOS_SISTEMA
                    (NOME, EMAIL, SENHA_HASH, TIPO_ACESSO, CABO_ID, ATIVO)
                VALUES
                    (:1, :2, :3, :4, :5, :6)
            """, (
                nome,
                email,
                senha_hash,
                tipo_acesso,
                int(cabo_id) if cabo_id else None,
                ativo
            ))

            conexao.commit()
            registrar_auditoria(
            acao="CADASTRO_CONTATO",
            tabela_afetada="CONTATOS_CAMPANHA",
            descricao=f"Apoiador {nome} cadastrado para liderança ID {cabo_id}.")

            flash("Usuário cadastrado com sucesso.", "success")
            return redirect(url_for("listar_usuarios"))

    except oracledb.Error as erro:
        conexao.rollback()
        print("ERRO CADASTRAR USUÁRIO:", erro)
        flash("Erro ao cadastrar usuário.", "danger")
        return redirect(url_for("home"))

    finally:
        cursor.close()
        conexao.close()

    return render_template("cadastrar_usuario.html", cabos=cabos)


@app.route("/listar-usuarios")
@login_required
@perfil_required("ADMIN")
def listar_usuarios():
    """
    Lista todos os usuários cadastrados no sistema.

    Exibe:
    - nome;
    - e-mail;
    - tipo de acesso;
    - liderança vinculada, quando existir;
    - status ativo/inativo;
    - data de cadastro.
    """

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("home"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT
                u.ID,
                u.NOME,
                u.EMAIL,
                u.TIPO_ACESSO,
                u.CABO_ID,
                c.NOME AS NOME_CABO,
                u.ATIVO,
                u.DATA_CADASTRO
            FROM USUARIOS_SISTEMA u
            LEFT JOIN CABOS_ELEITORAIS c
                ON c.ID = u.CABO_ID
            ORDER BY u.ID
        """)

        usuarios = cursor.fetchall()

    except oracledb.Error as erro:
        print("ERRO LISTAR USUÁRIOS:", erro)
        flash("Erro ao listar usuários.", "danger")
        return redirect(url_for("home"))

    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "listar_usuarios.html",
        usuarios=usuarios
    )

# ============================================================
# EDITAR USUÁRIO
# ============================================================

@app.route("/editar-usuario/<int:id>", methods=["GET", "POST"])
@login_required
@perfil_required("ADMIN")
def editar_usuario(id):
    """
    Edita dados de um usuário do sistema.

    Regras:
    - apenas ADMIN pode editar usuários;
    - e-mail não pode duplicar com outro usuário;
    - se nova senha for informada, atualiza o hash;
    - se não informar senha, mantém a senha atual.
    """

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar_usuarios"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT ID, NOME
            FROM CABOS_ELEITORAIS
            ORDER BY NOME
        """)
        cabos = cursor.fetchall()

        cursor.execute("""
            SELECT
                ID,
                NOME,
                EMAIL,
                TIPO_ACESSO,
                CABO_ID,
                ATIVO
            FROM USUARIOS_SISTEMA
            WHERE ID = :1
        """, (id,))
        usuario = cursor.fetchone()

        if not usuario:
            flash("Usuário não encontrado.", "warning")
            return redirect(url_for("listar_usuarios"))

        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            email = request.form.get("email", "").strip().lower()
            tipo_acesso = request.form.get("tipo_acesso", "").strip()
            cabo_id = request.form.get("cabo_id", "").strip()
            ativo = request.form.get("ativo", "S").strip()
            nova_senha = request.form.get("senha", "").strip()

            if not nome or not email or not tipo_acesso:
                flash("Preencha os campos obrigatórios.", "warning")
                return render_template("editar_usuario.html", usuario=usuario, cabos=cabos)

            cursor.execute("""
                SELECT ID
                FROM USUARIOS_SISTEMA
                WHERE LOWER(EMAIL) = LOWER(:1)
                  AND ID <> :2
            """, (email, id))
            email_existente = cursor.fetchone()

            if email_existente:
                flash("Já existe outro usuário com esse e-mail.", "warning")
                return render_template("editar_usuario.html", usuario=usuario, cabos=cabos)

            if tipo_acesso == "CABO" and not cabo_id:
                flash("Para usuário do tipo CABO, selecione a liderança.", "warning")
                return render_template("editar_usuario.html", usuario=usuario, cabos=cabos)

            if nova_senha:
                senha_hash = generate_password_hash(nova_senha)

                cursor.execute("""
                    UPDATE USUARIOS_SISTEMA
                    SET NOME = :1,
                        EMAIL = :2,
                        SENHA_HASH = :3,
                        TIPO_ACESSO = :4,
                        CABO_ID = :5,
                        ATIVO = :6
                    WHERE ID = :7
                """, (
                    nome,
                    email,
                    senha_hash,
                    tipo_acesso,
                    int(cabo_id) if cabo_id else None,
                    ativo,
                    id
                ))

            else:
                cursor.execute("""
                    UPDATE USUARIOS_SISTEMA
                    SET NOME = :1,
                        EMAIL = :2,
                        TIPO_ACESSO = :3,
                        CABO_ID = :4,
                        ATIVO = :5
                    WHERE ID = :6
                """, (
                    nome,
                    email,
                    tipo_acesso,
                    int(cabo_id) if cabo_id else None,
                    ativo,
                    id
                ))

            conexao.commit()

            flash("Usuário atualizado com sucesso.", "success")
            return redirect(url_for("listar_usuarios"))

    except oracledb.Error as erro:
        conexao.rollback()
        print("ERRO EDITAR USUÁRIO:", erro)
        flash("Erro ao editar usuário.", "danger")
        return redirect(url_for("listar_usuarios"))

    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "editar_usuario.html",
        usuario=usuario,
        cabos=cabos
    )


# ============================================================
# INATIVAR USUÁRIO
# ============================================================

@app.route("/inativar-usuario/<int:id>")
@login_required
@perfil_required("ADMIN")
def inativar_usuario(id):
    """
    Inativa um usuário do sistema.

    Observação:
    - não exclui fisicamente o usuário;
    - apenas altera ATIVO para 'N';
    - preserva histórico e segurança.
    """

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar_usuarios"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            UPDATE USUARIOS_SISTEMA
            SET ATIVO = 'N'
            WHERE ID = :1
        """, (id,))

        conexao.commit()

        flash("Usuário inativado com sucesso.", "success")

    except oracledb.Error as erro:
        conexao.rollback()
        print("ERRO INATIVAR USUÁRIO:", erro)
        flash("Erro ao inativar usuário.", "danger")

    finally:
        cursor.close()
        conexao.close()

    return redirect(url_for("listar_usuarios"))


# ============================================================
# ESQUECI MINHA SENHA
# ============================================================

@app.route("/esqueci-senha", methods=["GET", "POST"])
def esqueci_senha():
    """
    Solicita recuperação de senha.

    Fluxo:
    - usuário informa e-mail;
    - sistema verifica se existe usuário ativo;
    - gera token seguro;
    - salva token com expiração;
    - envia link de redefinição por e-mail.
    """

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()

        if not email:
            flash("Informe o e-mail cadastrado.", "warning")
            return render_template("esqueci_senha.html")

        conexao = conectar_oracle()

        if conexao is None:
            flash("Não foi possível conectar ao banco de dados.", "danger")
            return redirect(url_for("login"))

        cursor = conexao.cursor()

        try:
            cursor.execute("""
                SELECT ID, NOME, EMAIL
                FROM USUARIOS_SISTEMA
                WHERE LOWER(EMAIL) = LOWER(:1)
                  AND ATIVO = 'S'
            """, (email,))

            usuario = cursor.fetchone()

            if usuario:
                token = serializer.dumps(email, salt="recuperar-senha")
                data_expiracao = datetime.now() + timedelta(minutes=30)

                cursor.execute("""
                    INSERT INTO RECUPERACAO_SENHA
                        (USUARIO_ID, TOKEN, DATA_EXPIRACAO)
                    VALUES
                        (:1, :2, :3)
                """, (
                    usuario[0],
                    token,
                    data_expiracao
                ))

                conexao.commit()

                link = f"{BASE_URL}/redefinir-senha/{token}"

                email_enviado = enviar_email_recuperacao_senha(
                    destinatario=email,
                    nome_usuario=usuario[1],
                    link=link
                )

                if not email_enviado:
                    print("ERRO: token gerado, mas e-mail de recuperação não foi enviado.")

            flash(
                "Se o e-mail existir no sistema, um link de recuperação foi enviado.",
                "info"
            )

            return redirect(url_for("login"))

        except oracledb.Error as erro:
            conexao.rollback()
            print("ERRO ESQUECI SENHA:", erro)
            flash("Erro ao solicitar recuperação de senha.", "danger")
            return redirect(url_for("login"))

        finally:
            cursor.close()
            conexao.close()

    return render_template("esqueci_senha.html")

# ============================================================
# REDEFINIR SENHA
# ============================================================

@app.route("/redefinir-senha/<token>", methods=["GET", "POST"])
def redefinir_senha(token):
    """
    Redefine a senha do usuário a partir de um token válido.

    Regras:
    - token precisa existir;
    - token não pode estar usado;
    - token não pode estar expirado;
    - nova senha precisa ter pelo menos 8 caracteres;
    - senha é salva com hash;
    - token é marcado como usado após a redefinição.
    """

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("login"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT
                r.ID,
                r.USUARIO_ID,
                r.TOKEN,
                r.DATA_EXPIRACAO,
                r.USADO,
                u.EMAIL
            FROM RECUPERACAO_SENHA r
            JOIN USUARIOS_SISTEMA u
                ON u.ID = r.USUARIO_ID
            WHERE r.TOKEN = :1
        """, (token,))

        recuperacao = cursor.fetchone()

        if not recuperacao:
            flash("Link de recuperação inválido.", "danger")
            return redirect(url_for("login"))

        recuperacao_id = recuperacao[0]
        usuario_id = recuperacao[1]
        data_expiracao = recuperacao[3]
        usado = recuperacao[4]

        if usado == "S":
            flash("Este link de recuperação já foi utilizado.", "warning")
            return redirect(url_for("login"))

        if data_expiracao < datetime.now():
            flash("Este link de recuperação expirou.", "warning")
            return redirect(url_for("login"))

        if request.method == "POST":
            nova_senha = request.form.get("nova_senha", "").strip()
            confirmar_senha = request.form.get("confirmar_senha", "").strip()

            if not nova_senha or not confirmar_senha:
                flash("Informe e confirme a nova senha.", "warning")
                return render_template("redefinir_senha.html")

            if nova_senha != confirmar_senha:
                flash("As senhas não conferem.", "warning")
                return render_template("redefinir_senha.html")

            if len(nova_senha) < 8:
                flash("A senha deve ter pelo menos 8 caracteres.", "warning")
                return render_template("redefinir_senha.html")

            senha_hash = generate_password_hash(nova_senha)

            cursor.execute("""
                UPDATE USUARIOS_SISTEMA
                SET SENHA_HASH = :1
                WHERE ID = :2
            """, (
                senha_hash,
                usuario_id
            ))

            cursor.execute("""
                UPDATE RECUPERACAO_SENHA
                SET USADO = 'S'
                WHERE ID = :1
            """, (recuperacao_id,))

            conexao.commit()

            flash(
                "Senha redefinida com sucesso. Faça login novamente.",
                "success"
            )

            return redirect(url_for("login"))

    except Exception as erro:
        conexao.rollback()
        print("ERRO AO REDEFINIR SENHA:", erro)
        flash("Não foi possível redefinir a senha.", "danger")
        return redirect(url_for("login"))

    finally:
        cursor.close()
        conexao.close()

    return render_template("redefinir_senha.html")

# ============================================================
# ENVIAR CONVITE
# ============================================================

@app.route("/enviar-convite/<int:cabo_id>", methods=["GET", "POST"])
@login_required
def enviar_convite(cabo_id):
    tipo_acesso = session.get("tipo_acesso")
    cabo_sessao = session.get("cabo_id")

    if tipo_acesso == "CABO" and cabo_sessao != cabo_id:
        flash("Você só pode enviar convites da sua própria liderança.", "danger")
        return redirect(url_for("home"))

    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()
    token = None

    try:
        cursor.execute("""
            SELECT ID, NOME, EMAIL
            FROM CABOS_ELEITORAIS
            WHERE ID = :1
        """, (cabo_id,))
        cabo = cursor.fetchone()

        if not cabo:
            flash("Liderança não encontrada.", "warning")
            return redirect(url_for("listar"))

        if request.method == "GET":
            return render_template("enviar_convite.html", cabo=cabo)

        email = request.form.get("email", "").strip().lower()
        telefone = request.form.get("telefone", "").strip()
        forma_envio = request.form.get("forma_envio", "EMAIL").strip()

        if forma_envio in ("EMAIL", "AMBOS") and not email:
            flash("Informe o e-mail do apoiador.", "warning")
            return render_template("enviar_convite.html", cabo=cabo)

        if forma_envio in ("WHATSAPP", "AMBOS") and not telefone:
            flash("Informe o telefone/WhatsApp do apoiador.", "warning")
            return render_template("enviar_convite.html", cabo=cabo)

        if email:
            cursor.execute("""
                SELECT ID
                FROM CONTATOS_CAMPANHA
                WHERE LOWER(EMAIL) = LOWER(:1)
            """, (email,))
            contato_existente = cursor.fetchone()

            if contato_existente:
                flash("Já existe um apoiador cadastrado com esse e-mail.", "warning")
                return render_template("enviar_convite.html", cabo=cabo)

            cursor.execute("""
                SELECT cc.ID, c.NOME
                FROM CONVITES_CONTATO cc
                JOIN CABOS_ELEITORAIS c
                    ON c.ID = cc.CABO_ID
                WHERE LOWER(cc.EMAIL) = LOWER(:1)
                  AND cc.STATUS = 'PENDENTE'
            """, (email,))
            convite_existente = cursor.fetchone()

            if convite_existente:
                flash("Já existe um convite pendente para este e-mail.", "warning")
                return render_template("enviar_convite.html", cabo=cabo)

        token = gerar_token_convite()

        cursor.execute("""
            INSERT INTO CONVITES_CONTATO
                (
                    EMAIL,
                    TELEFONE,
                    CABO_ID,
                    TOKEN,
                    STATUS,
                    STATUS_ENVIO,
                    DATA_EXPIRACAO
                )
            VALUES
                (:1, :2, :3, :4, 'PENDENTE', 'PENDENTE', SYSDATE + 7)
        """, (
            email,
            telefone,
            cabo_id,
            token
        ))

        conexao.commit()

        link_convite = f"{BASE_URL}/convite/{token}"

        mensagem_whatsapp = f"""
Olá! Você recebeu um convite para cadastro no Sistema de Lideranças.

Liderança responsável: {cabo[1]}

Acesse o link:
{link_convite}
"""

        telefone_limpo = "".join(filter(str.isdigit, telefone))
        whatsapp_link = None

        if telefone_limpo:
            if not telefone_limpo.startswith("55"):
                telefone_limpo = "55" + telefone_limpo

            whatsapp_link = f"https://wa.me/{telefone_limpo}?text={quote(mensagem_whatsapp)}"

        email_enviado = False

        if forma_envio in ("EMAIL", "AMBOS"):
            email_enviado = enviar_email_convite(email, cabo[1], link_convite)

        if forma_envio == "WHATSAPP":
            cursor.execute("""
                UPDATE CONVITES_CONTATO
                SET STATUS_ENVIO = 'WHATSAPP',
                    ERRO_ENVIO = NULL,
                    DATA_ENVIO = SYSDATE
                WHERE TOKEN = :1
            """, (token,))
            conexao.commit()
            flash("Convite gerado para envio via WhatsApp.", "success")

        elif email_enviado:
            cursor.execute("""
                UPDATE CONVITES_CONTATO
                SET STATUS_ENVIO = 'ENVIADO',
                    ERRO_ENVIO = NULL,
                    DATA_ENVIO = SYSDATE
                WHERE TOKEN = :1
            """, (token,))
            conexao.commit()
            flash("Convite enviado com sucesso.", "success")

        else:
            cursor.execute("""
                UPDATE CONVITES_CONTATO
                SET STATUS_ENVIO = 'ERRO',
                    ERRO_ENVIO = 'Falha ao enviar e-mail'
                WHERE TOKEN = :1
            """, (token,))
            conexao.commit()
            flash("Convite gerado, mas houve falha no envio do e-mail.", "warning")

        return render_template(
            "convite_enviado.html",
            whatsapp_link=whatsapp_link,
            email=email
        )

    except Exception as erro:
        if token:
            try:
                cursor.execute("""
                    UPDATE CONVITES_CONTATO
                    SET STATUS_ENVIO = 'ERRO',
                        ERRO_ENVIO = :1
                    WHERE TOKEN = :2
                """, (
                    str(erro)[:500],
                    token
                ))
                conexao.commit()
            except Exception:
                pass

        print("ERRO AO ENVIAR CONVITE:", erro)
        flash("Erro ao enviar convite.", "danger")
        return redirect(url_for("listar"))

    finally:
        cursor.close()
        conexao.close()
        
    #return render_template("enviar_convite.html", cabo=cabo)

# ============================================================
# CADASTRO POR CONVITE
# ============================================================

@app.route("/convite/<token>", methods=["GET", "POST"])
def cadastro_por_convite(token):
    """
    Permite que o apoiador conclua o cadastro usando o link de convite.

    Regras:
    - token precisa existir;
    - convite precisa estar PENDENTE;
    - convite não pode estar expirado;
    - liderança vinculada precisa existir;
    - não permite duplicidade de nome para a mesma liderança;
    - ao concluir, cria o apoiador e marca o convite como USADO.
    """

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("login"))

    cursor = conexao.cursor()

    try:
        # Busca o convite pelo token recebido na URL.
        cursor.execute("""
            SELECT
                ID,
                EMAIL,
                CABO_ID,
                STATUS,
                DATA_EXPIRACAO
            FROM CONVITES_CONTATO
            WHERE TOKEN = :1
        """, (token,))

        convite = cursor.fetchone()

        if not convite:
            flash("Convite inválido.", "danger")
            return redirect(url_for("login"))

        convite_id = convite[0]
        email = convite[1]
        cabo_id = convite[2]
        status = convite[3]
        data_expiracao = convite[4]

        # Se o convite estiver vencido, marca como EXPIRADO.
        if data_expiracao and data_expiracao < datetime.now():
            cursor.execute("""
                UPDATE CONVITES_CONTATO
                SET STATUS = 'EXPIRADO'
                WHERE ID = :1
                  AND STATUS = 'PENDENTE'
            """, (convite_id,))

            conexao.commit()

            flash("Este convite está expirado.", "warning")
            return redirect(url_for("login"))

        # Apenas convites pendentes podem ser usados.
        if status != "PENDENTE":
            flash(
                "Este convite já foi utilizado ou não está mais disponível.",
                "warning"
            )
            return redirect(url_for("login"))

        # Busca a liderança responsável pelo convite.
        cursor.execute("""
            SELECT NOME
            FROM CABOS_ELEITORAIS
            WHERE ID = :1
        """, (cabo_id,))

        cabo = cursor.fetchone()

        if not cabo:
            flash("Liderança vinculada ao convite não encontrada.", "danger")
            return redirect(url_for("login"))

        nome_cabo = cabo[0]

        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            telefone = request.form.get("telefone", "").strip()
            endereco = request.form.get("endereco", "").strip()
            regiao = request.form.get("regiao_administrativa", "").strip()
            data_nascimento = converter_data(
                request.form.get("data_nascimento", "").strip()
            )
            cep = request.form.get("cep", "").strip()
            consentiu = request.form.get("consentiu_contato", "N")
            observacao = request.form.get("observacao", "").strip()

            if not nome:
                flash("O nome é obrigatório.", "warning")
                return render_template(
                    "cadastro_por_convite.html",
                    email=email,
                    nome_cabo=nome_cabo
                )

            # Impede duplicidade de nome na mesma liderança.
            cursor.execute("""
                SELECT ID
                FROM CONTATOS_CAMPANHA
                WHERE UPPER(NOME) = UPPER(:1)
                  AND CABO_ID = :2
            """, (
                nome,
                cabo_id
            ))

            registro_existente = cursor.fetchone()

            if registro_existente:
                flash(
                    "Já existe um apoiador cadastrado para esta liderança com esse nome.",
                    "warning"
                )
                return render_template(
                    "cadastro_por_convite.html",
                    email=email,
                    nome_cabo=nome_cabo
                )

            # Insere o apoiador.
            cursor.execute("""
                INSERT INTO CONTATOS_CAMPANHA
                    (
                        NOME,
                        EMAIL,
                        TELEFONE,
                        ENDERECO,
                        REGIAO_ADMINISTRATIVA,
                        CABO_ID,
                        CONSENTIU_CONTATO,
                        OBSERVACAO,
                        DATA_NASCIMENTO,
                        CEP
                    )
                VALUES
                    (:1, :2, :3, :4, :5, :6, :7, :8, :9, :10)
            """, (
                nome,
                email,
                telefone,
                endereco,
                regiao,
                cabo_id,
                consentiu,
                observacao,
                data_nascimento,
                cep
            ))

            # Marca o convite como utilizado.
            cursor.execute("""
                UPDATE CONVITES_CONTATO
                SET STATUS = 'USADO',
                    DATA_USO = SYSDATE
                WHERE ID = :1
            """, (convite_id,))

            conexao.commit()

            flash("Cadastro concluído com sucesso.", "success")
            return redirect(url_for("login"))

    except oracledb.Error as erro:
        conexao.rollback()
        print("ERRO CADASTRO POR CONVITE:", erro)
        flash("Erro ao processar cadastro por convite.", "danger")
        return redirect(url_for("login"))

    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "cadastro_por_convite.html",
        email=email,
        nome_cabo=nome_cabo
    )

# ============================================================
# LISTAR CONVITES
# ============================================================

@app.route("/listar-convites")
@login_required
def listar_convites():
    """
    Lista os convites enviados pelo sistema.

    Funcionalidades:
    - filtro por e-mail;
    - filtro por status do convite;
    - filtro por status de envio;
    - filtro por liderança;
    - usuário CABO visualiza apenas seus próprios convites.
    """

    termo = request.args.get("busca", "").strip()
    status = request.args.get("status", "").strip()
    status_envio = request.args.get("status_envio", "").strip()
    cabo_id = request.args.get("cabo_id", "").strip()

    tipo_acesso = session.get("tipo_acesso")
    cabo_sessao = session.get("cabo_id")

    registros = []
    cabos = []

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return render_template(
            "listar_convites.html",
            registros=registros,
            termo=termo,
            status=status,
            status_envio=status_envio,
            cabo_id=cabo_id,
            cabos=cabos
        )

    cursor = conexao.cursor()

    try:
        # Carrega lideranças para o filtro.
        # CABO vê apenas sua própria liderança.
        if tipo_acesso == "CABO":
            cursor.execute("""
                SELECT ID, NOME
                FROM CABOS_ELEITORAIS
                WHERE ID = :1
                ORDER BY NOME
            """, (cabo_sessao,))
        else:
            cursor.execute("""
                SELECT ID, NOME
                FROM CABOS_ELEITORAIS
                ORDER BY NOME
            """)

        cabos = cursor.fetchall()

        # Consulta principal dos convites.
        sql = """
            SELECT
                cc.ID,
                cc.EMAIL,
                c.NOME,
                cc.STATUS,
                cc.STATUS_ENVIO,
                cc.DATA_ENVIO,
                cc.DATA_EXPIRACAO,
                cc.DATA_USO,
                cc.ERRO_ENVIO
            FROM CONVITES_CONTATO cc
            JOIN CABOS_ELEITORAIS c
                ON c.ID = cc.CABO_ID
            WHERE 1=1
        """

        params = {}

        if termo:
            sql += """
                AND TRIM(UPPER(cc.EMAIL))
                    LIKE TRIM(UPPER(:busca))
            """
            params["busca"] = f"%{termo}%"

        if status:
            sql += " AND cc.STATUS = :status"
            params["status"] = status

        if status_envio:
            sql += " AND cc.STATUS_ENVIO = :status_envio"
            params["status_envio"] = status_envio

        if cabo_id:
            sql += " AND cc.CABO_ID = :cabo_id"
            params["cabo_id"] = int(cabo_id)

        if tipo_acesso == "CABO":
            sql += " AND cc.CABO_ID = :cabo_sessao"
            params["cabo_sessao"] = cabo_sessao

        sql += """
            ORDER BY cc.ID DESC
        """

        cursor.execute(sql, params)
        registros = cursor.fetchall()

    except oracledb.Error as erro:
        print("ERRO AO LISTAR CONVITES:", erro)
        flash("Erro ao listar convites.", "danger")

    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "listar_convites.html",
        registros=registros,
        termo=termo,
        status=status,
        status_envio=status_envio,
        cabo_id=cabo_id,
        cabos=cabos
    )


# ============================================================
# REENVIAR CONVITE
# ============================================================

@app.route("/reenviar-convite/<int:convite_id>")
@login_required
def reenviar_convite(convite_id):
    """
    Reenvia um convite já criado.

    Regras:
    - busca o convite pelo ID;
    - não reenvia convite usado ou cancelado;
    - envia novamente por e-mail;
    - atualiza STATUS_ENVIO e DATA_ENVIO.
    """

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar_convites"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT
                cc.EMAIL,
                cc.TOKEN,
                cc.STATUS,
                c.NOME
            FROM CONVITES_CONTATO cc
            JOIN CABOS_ELEITORAIS c
                ON c.ID = cc.CABO_ID
            WHERE cc.ID = :1
        """, (convite_id,))

        convite = cursor.fetchone()

        if not convite:
            flash("Convite não encontrado.", "warning")
            return redirect(url_for("listar_convites"))

        email = convite[0]
        token = convite[1]
        status = convite[2]
        nome_cabo = convite[3]

        if status in ("USADO", "CANCELADO"):
            flash("Este convite não pode ser reenviado.", "warning")
            return redirect(url_for("listar_convites"))

        link_convite = f"{BASE_URL}/convite/{token}"

        email_enviado = enviar_email_convite(
            email,
            nome_cabo,
            link_convite
        )

        if email_enviado:
            cursor.execute("""
                UPDATE CONVITES_CONTATO
                SET STATUS_ENVIO = 'ENVIADO',
                    ERRO_ENVIO = NULL,
                    DATA_ENVIO = SYSDATE
                WHERE ID = :1
            """, (convite_id,))

            conexao.commit()

            flash("Convite reenviado com sucesso.", "success")

        else:
            cursor.execute("""
                UPDATE CONVITES_CONTATO
                SET STATUS_ENVIO = 'ERRO',
                    ERRO_ENVIO = 'Falha ao reenviar e-mail'
                WHERE ID = :1
            """, (convite_id,))

            conexao.commit()

            flash(
                "Não foi possível reenviar o convite por e-mail.",
                "warning"
            )

    except Exception as erro:
        try:
            cursor.execute("""
                UPDATE CONVITES_CONTATO
                SET STATUS_ENVIO = 'ERRO',
                    ERRO_ENVIO = :1
                WHERE ID = :2
            """, (
                str(erro)[:500],
                convite_id
            ))
            conexao.commit()
        except Exception:
            pass

        print("ERRO REENVIAR CONVITE:", erro)
        flash("Erro ao reenviar convite.", "danger")

    finally:
        cursor.close()
        conexao.close()

    return redirect(url_for("listar_convites"))

# ============================================================
# CANCELAR CONVITE
# ============================================================

@app.route("/cancelar-convite/<int:convite_id>")
@login_required
@perfil_required("ADMIN", "CHEFE_GABINETE", "SECRETARIA")
def cancelar_convite(convite_id):
    """
    Cancela um convite pendente.

    Regras:
    - não cancela convite já usado;
    - não cancela novamente convite já cancelado;
    - marca STATUS como CANCELADO;
    - mantém histórico no banco.
    """

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar_convites"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT ID, STATUS
            FROM CONVITES_CONTATO
            WHERE ID = :1
        """, (convite_id,))

        convite = cursor.fetchone()

        if not convite:
            flash("Convite não encontrado.", "warning")
            return redirect(url_for("listar_convites"))

        status_atual = convite[1]

        if status_atual == "USADO":
            flash("Não é possível cancelar um convite que já foi utilizado.", "warning")
            return redirect(url_for("listar_convites"))

        if status_atual == "CANCELADO":
            flash("Este convite já está cancelado.", "info")
            return redirect(url_for("listar_convites"))

        cursor.execute("""
            UPDATE CONVITES_CONTATO
            SET STATUS = 'CANCELADO',
                ERRO_ENVIO = 'Convite cancelado pelo usuário'
            WHERE ID = :1
        """, (convite_id,))

        conexao.commit()

        flash("Convite cancelado com sucesso.", "success")

    except oracledb.Error as erro:
        conexao.rollback()
        print("ERRO AO CANCELAR CONVITE:", erro)
        flash("Erro ao cancelar convite.", "danger")

    finally:
        cursor.close()
        conexao.close()

    return redirect(url_for("listar_convites"))


# ============================================================
# EXPORTAR CONVITES PARA EXCEL
# ============================================================

@app.route("/exportar-convites-excel")
@login_required
def exportar_convites_excel():
    """
    Exporta os convites enviados para Excel.

    Funcionalidades:
    - respeita filtros da tela de convites;
    - filtra por e-mail, status, status de envio e liderança;
    - usuário CABO exporta apenas seus próprios convites.
    """

    termo = request.args.get("busca", "").strip()
    status = request.args.get("status", "").strip()
    status_envio = request.args.get("status_envio", "").strip()
    cabo_id = request.args.get("cabo_id", "").strip()

    tipo_acesso = session.get("tipo_acesso")
    cabo_sessao = session.get("cabo_id")

    conexao = conectar_oracle()

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar_convites"))

    cursor = conexao.cursor()

    try:
        sql = """
            SELECT
                cc.ID,
                cc.EMAIL,
                c.NOME AS LIDERANCA,
                cc.STATUS,
                cc.STATUS_ENVIO,
                cc.DATA_ENVIO,
                cc.DATA_EXPIRACAO,
                cc.DATA_USO,
                cc.ERRO_ENVIO
            FROM CONVITES_CONTATO cc
            JOIN CABOS_ELEITORAIS c
                ON c.ID = cc.CABO_ID
            WHERE 1=1
        """

        params = {}

        if termo:
            sql += """
                AND TRIM(UPPER(cc.EMAIL))
                    LIKE TRIM(UPPER(:busca))
            """
            params["busca"] = f"%{termo}%"

        if status:
            sql += " AND cc.STATUS = :status"
            params["status"] = status

        if status_envio:
            sql += " AND cc.STATUS_ENVIO = :status_envio"
            params["status_envio"] = status_envio

        if cabo_id:
            sql += " AND cc.CABO_ID = :cabo_id"
            params["cabo_id"] = int(cabo_id)

        if tipo_acesso == "CABO":
            sql += " AND cc.CABO_ID = :cabo_sessao"
            params["cabo_sessao"] = cabo_sessao

        sql += """
            ORDER BY cc.ID DESC
        """

        cursor.execute(sql, params)
        registros = cursor.fetchall()

    except oracledb.Error as erro:
        print("ERRO AO EXPORTAR CONVITES:", erro)
        flash("Erro ao exportar convites.", "danger")
        return redirect(url_for("listar_convites"))

    finally:
        cursor.close()
        conexao.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Convites"

    sheet.append([
        "ID",
        "E-mail",
        "Liderança",
        "Status do Convite",
        "Status do Envio",
        "Data de Envio",
        "Data de Expiração",
        "Data de Uso",
        "Erro de Envio"
    ])

    for registro in registros:
        linha = list(registro)

        if linha[5]:
            linha[5] = linha[5].strftime("%d/%m/%Y %H:%M")

        if linha[6]:
            linha[6] = linha[6].strftime("%d/%m/%Y %H:%M")

        if linha[7]:
            linha[7] = linha[7].strftime("%d/%m/%Y %H:%M")

        sheet.append(linha)

    arquivo_excel = BytesIO()
    workbook.save(arquivo_excel)
    arquivo_excel.seek(0)

    return send_file(
        arquivo_excel,
        as_attachment=True,
        download_name="convites_enviados.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ============================================================
# REGISTRAR AUDITORIA REALIZADAS NO SISTEMA
# ============================================================
# ============================================================

def registrar_auditoria(acao, tabela_afetada=None, registro_id=None, descricao=None):
    """
    Registra ações importantes realizadas no sistema.
    """

    try:
        conexao = conectar_oracle()
        if conexao is None:
            return

        cursor = conexao.cursor()

        usuario_id = session.get("usuario_id")
        usuario_nome = session.get("usuario_nome")
        tipo_acesso = session.get("tipo_acesso")
        ip_acesso = request.headers.get("X-Forwarded-For", request.remote_addr)

        cursor.execute("""
            INSERT INTO AUDITORIA_LOGS
                (
                    USUARIO_ID,
                    USUARIO_NOME,
                    TIPO_ACESSO,
                    ACAO,
                    TABELA_AFETADA,
                    REGISTRO_ID,
                    DESCRICAO,
                    IP_ACESSO
                )
            VALUES
                (:1, :2, :3, :4, :5, :6, :7, :8)
        """, (
            usuario_id,
            usuario_nome,
            tipo_acesso,
            acao,
            tabela_afetada,
            registro_id,
            descricao,
            ip_acesso
        ))

        conexao.commit()
        cursor.close()
        conexao.close()

    except Exception as erro:
        print("ERRO AO REGISTRAR AUDITORIA:", erro)
        
# ============================================================
# Logs de Auditoria para o ADMIN consultar tudo
# ============================================================

@app.route("/auditoria")
@login_required
@perfil_required("ADMIN")
def auditoria():
    termo = request.args.get("busca", "").strip()
    acao = request.args.get("acao", "").strip()

    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("home"))

    cursor = conexao.cursor()

    try:
        sql = """
            SELECT
                ID,
                USUARIO_NOME,
                TIPO_ACESSO,
                ACAO,
                TABELA_AFETADA,
                REGISTRO_ID,
                DESCRICAO,
                IP_ACESSO,
                DATA_LOG
            FROM AUDITORIA_LOGS
            WHERE 1=1
        """

        params = {}

        if termo:
            sql += """
                AND (
                    UPPER(USUARIO_NOME) LIKE UPPER(:busca)
                    OR UPPER(DESCRICAO) LIKE UPPER(:busca)
                    OR UPPER(IP_ACESSO) LIKE UPPER(:busca)
                )
            """
            params["busca"] = f"%{termo}%"

        if acao:
            sql += " AND ACAO = :acao"
            params["acao"] = acao

        sql += " ORDER BY ID DESC"

        cursor.execute(sql, params)
        logs = cursor.fetchall()

        cursor.execute("""
            SELECT DISTINCT ACAO
            FROM AUDITORIA_LOGS
            ORDER BY ACAO
        """)
        acoes = [r[0] for r in cursor.fetchall()]

    except oracledb.Error as erro:
        print("ERRO AUDITORIA:", erro)
        flash("Erro ao carregar logs de auditoria.", "danger")
        return redirect(url_for("home"))

    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "auditoria.html",
        logs=logs,
        termo=termo,
        acao=acao,
        acoes=acoes
    )

# ============================================================
# EXECUÇÃO LOCAL
# ============================================================

if __name__ == "__main__":
    """
    Executa o Flask localmente.

    No Render, quem executa é o Gunicorn:
    gunicorn app:app --timeout 120
    """

    debug = os.getenv("FLASK_DEBUG", "False").lower() == "true"
    app.run(debug=debug)

