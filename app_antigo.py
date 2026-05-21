import os
import secrets
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from io import BytesIO
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
from dotenv import load_dotenv
import oracledb
from openpyxl import Workbook
from db import conectar_oracle


def converter_data(data_str):
    if not data_str:
        return None
    try:
        return datetime.strptime(data_str, "%Y-%m-%d").date()
    except ValueError:
        return None


def gerar_token_convite():
    return secrets.token_urlsafe(32)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))

SMTP_HOST = os.getenv("SMTP_HOST")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
MAIL_FROM = os.getenv("MAIL_FROM")
BASE_URL = os.getenv("BASE_URL", "http://127.0.0.1:5000")

ADMIN_USER = os.getenv("ADMIN_USER")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "chave_padrao")

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("usuario_logado"):
            flash("Faça login para acessar o sistema.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

@app.route("/")
def home():
    regiao = request.args.get("regiao", "").strip()

    conexao = conectar_oracle()
    total_cabos = 0
    total_contatos = 0
    ranking_cabos = []
    regioes = []

    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return render_template(
            "dashboard.html",
            total_cabos=0,
            total_contatos=0,
            ranking_cabos=[],
            regiao=regiao,
            regioes=[]
        )

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT DISTINCT REGIAO_ADMINISTRATIVA
            FROM CABOS_ELEITORAIS
            WHERE REGIAO_ADMINISTRATIVA IS NOT NULL
            ORDER BY REGIAO_ADMINISTRATIVA
        """)
        regioes = [r[0] for r in cursor.fetchall()]

        if regiao:
            cursor.execute("""
                SELECT COUNT(*)
                FROM CABOS_ELEITORAIS
                WHERE REGIAO_ADMINISTRATIVA = :1
            """, (regiao,))
        else:
            cursor.execute("SELECT COUNT(*) FROM CABOS_ELEITORAIS")
        total_cabos = cursor.fetchone()[0]

        if regiao:
            cursor.execute("""
                SELECT COUNT(ct.ID)
                FROM CONTATOS_CAMPANHA ct
                JOIN CABOS_ELEITORAIS c ON c.ID = ct.CABO_ID
                WHERE c.REGIAO_ADMINISTRATIVA = :1
            """, (regiao,))
        else:
            cursor.execute("SELECT COUNT(*) FROM CONTATOS_CAMPANHA")
        total_contatos = cursor.fetchone()[0]

        sql = """
            SELECT
                c.NOME,
                c.REGIAO_ADMINISTRATIVA,
                COUNT(ct.ID) AS TOTAL_CONTATOS
            FROM CABOS_ELEITORAIS c
            LEFT JOIN CONTATOS_CAMPANHA ct ON ct.CABO_ID = c.ID
            WHERE 1=1
        """
        params = {}

        if regiao:
            sql += " AND c.REGIAO_ADMINISTRATIVA = :regiao"
            params["regiao"] = regiao

        sql += """
            GROUP BY c.NOME, c.REGIAO_ADMINISTRATIVA
            ORDER BY TOTAL_CONTATOS DESC, c.NOME
        """

        cursor.execute(sql, params)
        ranking_cabos = cursor.fetchall()

    except oracledb.Error as erro:
        flash(f"Erro ao carregar dashboard: {erro}", "danger")
    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "dashboard.html",
        total_cabos=total_cabos,
        total_contatos=total_contatos,
        ranking_cabos=ranking_cabos,
        regiao=regiao,
        regioes=regioes
    )

@app.route("/cadastrar", methods=["GET", "POST"])
@login_required
def cadastrar():
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        email = request.form.get("email", "").strip()
        telefone = request.form.get("telefone", "").strip()
        endereco = request.form.get("endereco", "").strip()
        regiao_administrativa = request.form.get("regiao_administrativa", "").strip()
        data_nascimento = converter_data(request.form.get("data_nascimento", "").strip())
        cep = request.form.get("cep", "").strip()

        if not nome:
            flash("O campo nome é obrigatório.", "warning")
            return render_template("cadastrar.html")

        conexao = conectar_oracle()
        if conexao is None:
            flash("Não foi possível conectar ao banco de dados.", "danger")
            return render_template("cadastrar.html")

        cursor = conexao.cursor()

        try:
            cursor.execute("""
                INSERT INTO CABOS_ELEITORAIS
                (NOME, EMAIL, TELEFONE, ENDERECO, REGIAO_ADMINISTRATIVA, DATA_NASCIMENTO, CEP)
                VALUES (:1, :2, :3, :4, :5, :6, :7)
            """, (nome, email, telefone, endereco, regiao_administrativa, data_nascimento, cep))
            conexao.commit()
            flash("Cabo cadastrado com sucesso.", "success")
            return redirect(url_for("listar"))
        except oracledb.Error as erro:
            print("ERRO AO CADASTRAR CABO:", erro)
            flash(f"Erro ao cadastrar cabo: {erro}", "danger")
        finally:
            cursor.close()
            conexao.close()

    return render_template("cadastrar.html")

@app.route("/listar")
@login_required
def listar():
    termo = request.args.get("busca", "").strip()
    regiao = request.args.get("regiao", "").strip()

    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return render_template("listar.html", registros=[], termo=termo, regiao=regiao, regioes=[])

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT DISTINCT REGIAO_ADMINISTRATIVA
            FROM CABOS_ELEITORAIS
            WHERE REGIAO_ADMINISTRATIVA IS NOT NULL
            ORDER BY REGIAO_ADMINISTRATIVA
        """)
        regioes = [r[0] for r in cursor.fetchall()]

        sql = """
            SELECT
                c.ID,
                c.NOME,
                c.EMAIL,
                c.TELEFONE,
                c.ENDERECO,
                c.REGIAO_ADMINISTRATIVA,
                c.DATA_NASCIMENTO,
                c.CEP,
                COUNT(ct.ID) AS TOTAL_CONTATOS,
                SUM(CASE WHEN ct.CONSENTIU_CONTATO = 'S' THEN 1 ELSE 0 END) AS TOTAL_CONSENTIU,
                SUM(CASE WHEN ct.CONSENTIU_CONTATO = 'N' THEN 1 ELSE 0 END) AS TOTAL_NAO_CONSENTIU
            FROM CABOS_ELEITORAIS c
            LEFT JOIN CONTATOS_CAMPANHA ct ON ct.CABO_ID = c.ID
            WHERE 1=1
        """
        params = {}

        if termo:
            sql += " AND TRIM(UPPER(c.NOME)) LIKE TRIM(UPPER(:busca))"
            params["busca"] = f"%{termo}%"

        if regiao:
            sql += " AND c.REGIAO_ADMINISTRATIVA = :regiao"
            params["regiao"] = regiao

        sql += """
            GROUP BY
                c.ID,
                c.NOME,
                c.EMAIL,
                c.TELEFONE,
                c.ENDERECO,
                c.REGIAO_ADMINISTRATIVA,
                c.DATA_NASCIMENTO,
                c.CEP
            ORDER BY c.ID
        """

        cursor.execute(sql, params)
        registros = cursor.fetchall()
        print("REGISTROS LISTAR:", registros)

    except oracledb.Error as erro:
        flash(f"Erro ao listar dados: {erro}", "danger")
        print("ERRO NA LISTAGEM DE CABOS:", erro)
        registros = []
        regioes = []
    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "listar.html",
        registros=registros,
        termo=termo,
        regiao=regiao,
        regioes=regioes
    )

@app.route("/exportar-excel")
def exportar_excel():
    termo = request.args.get("busca", "").strip()
    regiao = request.args.get("regiao", "").strip()

    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()

    try:
        sql = """
            SELECT
                c.ID,
                c.NOME,
                c.EMAIL,
                c.TELEFONE,
                c.ENDERECO,
                c.REGIAO_ADMINISTRATIVA,
                c.DATA_NASCIMENTO,
                c.CEP,
                COUNT(ct.ID) AS TOTAL_CONTATOS
            FROM CABOS_ELEITORAIS c
            LEFT JOIN CONTATOS_CAMPANHA ct ON ct.CABO_ID = c.ID
            WHERE 1=1
        """
        params = {}

        if termo:
            sql += " AND TRIM(UPPER(c.NOME)) LIKE TRIM(UPPER(:busca))"
            params["busca"] = f"%{termo}%"

        if regiao:
            sql += " AND c.REGIAO_ADMINISTRATIVA = :regiao"
            params["regiao"] = regiao

        sql += """
            GROUP BY
                c.ID,
                c.NOME,
                c.EMAIL,
                c.TELEFONE,
                c.ENDERECO,
                c.REGIAO_ADMINISTRATIVA,
                c.DATA_NASCIMENTO,
                c.CEP
            ORDER BY c.ID
        """

        cursor.execute(sql, params)
        registros = cursor.fetchall()

    except oracledb.Error as erro:
        flash(f"Erro ao exportar dados: {erro}", "danger")
        print("ERRO AO EXPORTAR CABOS:", erro)
        return redirect(url_for("listar"))
    finally:
        cursor.close()
        conexao.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cabos Eleitorais"

    sheet.append([
        "ID",
        "Nome",
        "Email",
        "Telefone",
        "Endereço",
        "Região Administrativa",
        "Data de Nascimento",
        "CEP",
        "Total de Contatos"
    ])

    for registro in registros:
        linha = list(registro)

        if linha[6]:
            linha[6] = linha[6].strftime("%d/%m/%Y")

        sheet.append(linha)

    arquivo_excel = BytesIO()
    workbook.save(arquivo_excel)
    arquivo_excel.seek(0)

    nome_arquivo = "cabos_eleitorais_filtrados.xlsx" if termo or regiao else "cabos_eleitorais.xlsx"

    return send_file(
        arquivo_excel,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/editar/<int:id>", methods=["GET", "POST"])
@login_required
def editar(id):
    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        email = request.form.get("email", "").strip()
        telefone = request.form.get("telefone", "").strip()
        endereco = request.form.get("endereco", "").strip()
        regiao_administrativa = request.form.get("regiao_administrativa", "").strip()
        data_nascimento = converter_data(request.form.get("data_nascimento", "").strip())
        cep = request.form.get("cep", "").strip()

        try:
            cursor.execute("""
                UPDATE CABOS_ELEITORAIS
                SET NOME = :1,
                    EMAIL = :2,
                    TELEFONE = :3,
                    ENDERECO = :4,
                    REGIAO_ADMINISTRATIVA = :5,
                    DATA_NASCIMENTO = :6,
                    CEP = :7
                WHERE ID = :8
            """, (nome, email, telefone, endereco, regiao_administrativa, data_nascimento, cep, id))
            conexao.commit()
            flash("Registro atualizado com sucesso.", "success")
            return redirect(url_for("listar"))
        except oracledb.Error as erro:
            print("ERRO AO EDITAR CABO:", erro)
            flash(f"Erro ao atualizar cabo: {erro}", "danger")
        finally:
            cursor.close()
            conexao.close()

    try:
        cursor.execute("""
            SELECT ID, NOME, EMAIL, TELEFONE, ENDERECO, REGIAO_ADMINISTRATIVA, DATA_NASCIMENTO, CEP
            FROM CABOS_ELEITORAIS
            WHERE ID = :1
        """, (id,))
        registro = cursor.fetchone()
    except oracledb.Error as erro:
        print("ERRO AO BUSCAR CABO:", erro)
        flash(f"Erro ao buscar registro: {erro}", "danger")
        registro = None
    finally:
        cursor.close()
        conexao.close()

    if not registro:
        flash("Registro não encontrado.", "warning")
        return redirect(url_for("listar"))

    return render_template("editar.html", registro=registro)

@app.route("/excluir/<int:id>")
@login_required
def excluir(id):
    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()

    try:
        cursor.execute("DELETE FROM CABOS_ELEITORAIS WHERE ID = :1", (id,))
        conexao.commit()
        flash("Registro excluído com sucesso.", "success")
    except oracledb.Error as erro:
        flash(f"Erro ao excluir registro: {erro}", "danger")
    finally:
        cursor.close()
        conexao.close()

    return redirect(url_for("listar"))

@app.route("/cadastrar-contato", methods=["GET", "POST"])
@login_required
def cadastrar_contato():
    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("home"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT ID, NOME
            FROM CABOS_ELEITORAIS
            ORDER BY NOME
        """)
        cabos = cursor.fetchall()
    except oracledb.Error as erro:
        cursor.close()
        conexao.close()
        flash(f"Erro ao carregar cabos: {erro}", "danger")
        return redirect(url_for("home"))

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        telefone = request.form.get("telefone", "").strip()
        endereco = request.form.get("endereco", "").strip()
        regiao = request.form.get("regiao_administrativa", "").strip()
        cabo_id = request.form.get("cabo_id", "").strip()
        consentiu = request.form.get("consentiu_contato", "N")
        observacao = request.form.get("observacao", "").strip()
        data_nascimento = request.form.get("data_nascimento", "").strip()
        cep = request.form.get("cep", "").strip()

        if not nome or not cabo_id:
            flash("Nome e cabo responsável são obrigatórios.", "warning")
            cursor.close()
            conexao.close()
            return render_template("cadastrar_contato.html", cabos=cabos)

        try:
            if data_nascimento:
                cursor.execute("""
                    INSERT INTO CONTATOS_CAMPANHA
                    (NOME, TELEFONE, ENDERECO, REGIAO_ADMINISTRATIVA, CABO_ID, CONSENTIU_CONTATO, OBSERVACAO, DATA_NASCIMENTO, CEP)
                    VALUES (:1, :2, :3, :4, :5, :6, :7, TO_DATE(:8, 'YYYY-MM-DD'), :9)
                """, (nome, telefone, endereco, regiao, int(cabo_id), consentiu, observacao, data_nascimento, cep))
            else:
                cursor.execute("""
                    INSERT INTO CONTATOS_CAMPANHA
                    (NOME, TELEFONE, ENDERECO, REGIAO_ADMINISTRATIVA, CABO_ID, CONSENTIU_CONTATO, OBSERVACAO, DATA_NASCIMENTO, CEP)
                    VALUES (:1, :2, :3, :4, :5, :6, :7, NULL, :8)
                """, (nome, telefone, endereco, regiao, int(cabo_id), consentiu, observacao, cep))
            conexao.commit()
            flash("Contato cadastrado com sucesso.", "success")
            return redirect(url_for("listar_contatos"))
        except oracledb.Error as erro:
            flash(f"Erro ao cadastrar contato: {erro}", "danger")
        finally:
            cursor.close()
            conexao.close()

    cursor.close()
    conexao.close()
    return render_template("cadastrar_contato.html", cabos=cabos)

@app.route("/listar-contatos")
@login_required
def listar_contatos():
    termo = request.args.get("busca", "").strip()
    cabo_id = request.args.get("cabo_id", "").strip()
    regiao = request.args.get("regiao", "").strip()

    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return render_template(
            "listar_contatos.html",
            registros=[],
            termo=termo,
            cabo_id=cabo_id,
            regiao=regiao,
            cabos=[],
            regioes=[]
        )

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT ID, NOME
            FROM CABOS_ELEITORAIS
            ORDER BY NOME
        """)
        cabos = cursor.fetchall()

        cursor.execute("""
            SELECT DISTINCT REGIAO_ADMINISTRATIVA
            FROM CONTATOS_CAMPANHA
            WHERE REGIAO_ADMINISTRATIVA IS NOT NULL
            ORDER BY REGIAO_ADMINISTRATIVA
        """)
        regioes = [r[0] for r in cursor.fetchall()]

        sql = """
            SELECT
                ct.ID,                    -- item[0]
                ct.NOME,                  -- item[1]
                ct.TELEFONE,              -- item[2]
                ct.REGIAO_ADMINISTRATIVA, -- item[3]
                c.NOME AS CABO_NOME,      -- item[4]
                ct.DATA_NASCIMENTO,       -- item[5]
                ct.CEP,                   -- item[6]
                ct.CONSENTIU_CONTATO      -- item[7]
            FROM CONTATOS_CAMPANHA ct
            JOIN CABOS_ELEITORAIS c ON c.ID = ct.CABO_ID
            WHERE 1=1
        """
        params = {}

        if termo:
            sql += " AND UPPER(ct.NOME) LIKE :busca"
            params["busca"] = f"%{termo.upper()}%"

        if cabo_id:
            sql += " AND ct.CABO_ID = :cabo_id"
            params["cabo_id"] = int(cabo_id)

        if regiao:
            sql += " AND ct.REGIAO_ADMINISTRATIVA = :regiao"
            params["regiao"] = regiao

        sql += " ORDER BY ct.ID"

        cursor.execute(sql, params)
        registros = cursor.fetchall()

    except oracledb.Error as erro:
        flash(f"Erro ao listar contatos: {erro}", "danger")
        print("ERRO LISTAR CONTATOS:", erro)
        registros = []
        cabos = []
        regioes = []
    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "listar_contatos.html",
        registros=registros,
        termo=termo,
        cabo_id=cabo_id,
        regiao=regiao,
        cabos=cabos,
        regioes=regioes
    )
@app.route("/exportar-contatos-excel")
def exportar_contatos_excel():
    termo = request.args.get("busca", "").strip()
    cabo_id = request.args.get("cabo_id", "").strip()
    regiao = request.args.get("regiao", "").strip()

    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar_contatos"))

    cursor = conexao.cursor()

    try:
        sql = """
            SELECT
                ct.ID,
                ct.NOME,
                ct.TELEFONE,
                ct.ENDERECO,
                ct.REGIAO_ADMINISTRATIVA,
                c.NOME AS CABO_NOME,
                ct.CONSENTIU_CONTATO,
                ct.OBSERVACAO,
                ct.DATA_CADASTRO,
                ct.DATA_NASCIMENTO,
                ct.CEP
            FROM CONTATOS_CAMPANHA ct
            JOIN CABOS_ELEITORAIS c ON c.ID = ct.CABO_ID
            WHERE 1=1
        """
        params = {}

        if termo:
            sql += " AND UPPER(ct.NOME) LIKE :busca"
            params["busca"] = f"%{termo.upper()}%"

        if cabo_id:
            sql += " AND ct.CABO_ID = :cabo_id"
            params["cabo_id"] = int(cabo_id)

        if regiao:
            sql += " AND ct.REGIAO_ADMINISTRATIVA = :regiao"
            params["regiao"] = regiao

        sql += " ORDER BY ct.ID"

        cursor.execute(sql, params)
        registros = cursor.fetchall()

    except oracledb.Error as erro:
        flash(f"Erro ao exportar contatos: {erro}", "danger")
        print("ERRO AO EXPORTAR CONTATOS:", erro)
        return redirect(url_for("listar_contatos"))
    finally:
        cursor.close()
        conexao.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contatos de Campanha"

    sheet.append([
        "ID",
        "Nome",
        "Telefone",
        "Endereço",
        "Região Administrativa",
        "Cabo Responsável",
        "Consentiu Contato",
        "Observação",
        "Data Cadastro",
        "Data Nascimento",
        "CEP"
    ])

    for registro in registros:
        linha = list(registro)

        if linha[8]:
            linha[8] = linha[8].strftime("%d/%m/%Y")

        if linha[9]:
            linha[9] = linha[9].strftime("%d/%m/%Y")

        sheet.append(linha)

    arquivo_excel = BytesIO()
    workbook.save(arquivo_excel)
    arquivo_excel.seek(0)

    nome_arquivo = "contatos_filtrados.xlsx" if (termo or cabo_id or regiao) else "contatos_campanha.xlsx"

    return send_file(
        arquivo_excel,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/editar-contato/<int:id>", methods=["GET", "POST"])
@login_required
def editar_contato(id):
    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar_contatos"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT ID, NOME
            FROM CABOS_ELEITORAIS
            ORDER BY NOME
        """)
        cabos = cursor.fetchall()

        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            telefone = request.form.get("telefone", "").strip()
            endereco = request.form.get("endereco", "").strip()
            regiao = request.form.get("regiao_administrativa", "").strip()
            cabo_id = request.form.get("cabo_id", "").strip()
            consentiu = request.form.get("consentiu_contato", "N")
            observacao = request.form.get("observacao", "").strip()
            data_nascimento = converter_data(request.form.get("data_nascimento", "").strip())
            cep = request.form.get("cep", "").strip()

            if not nome or not cabo_id:
                flash("Nome e cabo responsável são obrigatórios.", "warning")
            else:
                cursor.execute("""
                    UPDATE CONTATOS_CAMPANHA
                    SET NOME = :1,
                        TELEFONE = :2,
                        ENDERECO = :3,
                        REGIAO_ADMINISTRATIVA = :4,
                        CABO_ID = :5,
                        CONSENTIU_CONTATO = :6,
                        OBSERVACAO = :7,
                        DATA_NASCIMENTO = :8,
                        CEP = :9
                    WHERE ID = :10
                """, (nome, telefone, endereco, regiao, int(cabo_id), consentiu, observacao, data_nascimento, cep, id))
                conexao.commit()
                flash("Contato atualizado com sucesso.", "success")
                return redirect(url_for("listar_contatos"))

        cursor.execute("""
            SELECT
                ID,
                NOME,
                TELEFONE,
                ENDERECO,
                REGIAO_ADMINISTRATIVA,
                CABO_ID,
                CONSENTIU_CONTATO,
                OBSERVACAO,
                DATA_NASCIMENTO,
                CEP
            FROM CONTATOS_CAMPANHA
            WHERE ID = :1
        """, (id,))
        contato = cursor.fetchone()

        if not contato:
            flash("Contato não encontrado.", "warning")
            return redirect(url_for("listar_contatos"))

    except oracledb.Error as erro:
        print("ERRO AO EDITAR CONTATO:", erro)
        flash(f"Erro ao carregar/atualizar contato: {erro}", "danger")
        return redirect(url_for("listar_contatos"))
    finally:
        cursor.close()
        conexao.close()

    return render_template("editar_contato.html", contato=contato, cabos=cabos)

@app.route("/excluir-contato/<int:id>")
@login_required
def excluir_contato(id):
    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar_contatos"))

    cursor = conexao.cursor()

    try:
        cursor.execute("DELETE FROM CONTATOS_CAMPANHA WHERE ID = :1", (id,))
        conexao.commit()
        flash("Contato excluído com sucesso.", "success")
    except oracledb.Error as erro:
        flash(f"Erro ao excluir contato: {erro}", "danger")
    finally:
        cursor.close()
        conexao.close()

    return redirect(url_for("listar_contatos"))

@app.route("/cabo/<int:cabo_id>/contatos")
def contatos_do_cabo(cabo_id):
    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT ID, NOME, REGIAO_ADMINISTRATIVA
            FROM CABOS_ELEITORAIS
            WHERE ID = :1
        """, (cabo_id,))
        cabo = cursor.fetchone()

        if not cabo:
            flash("Cabo não encontrado.", "warning")
            return redirect(url_for("listar"))

        cursor.execute("""
            SELECT
                ID,
                NOME,
                TELEFONE,
                ENDERECO,
                REGIAO_ADMINISTRATIVA,
                CONSENTIU_CONTATO,
                OBSERVACAO,
                DATA_CADASTRO,
                DATA_NASCIMENTO,
                CEP
            FROM CONTATOS_CAMPANHA
            WHERE CABO_ID = :1
            ORDER BY ID
        """, (cabo_id,))
        contatos = cursor.fetchall()

        total_contatos = len(contatos)

    except oracledb.Error as erro:
        flash(f"Erro ao carregar contatos do cabo: {erro}", "danger")
        return redirect(url_for("listar"))
    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "contatos_do_cabo.html",
        cabo=cabo,
        contatos=contatos,
        total_contatos=total_contatos
    )
    
@app.route("/cabo/<int:cabo_id>/exportar-excel")
def exportar_contatos_do_cabo(cabo_id):
    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT NOME
            FROM CABOS_ELEITORAIS
            WHERE ID = :1
        """, (cabo_id,))
        cabo = cursor.fetchone()

        if not cabo:
            flash("Cabo não encontrado.", "warning")
            return redirect(url_for("listar"))

        nome_cabo = cabo[0]

        cursor.execute("""
            SELECT
                ID,
                NOME,
                TELEFONE,
                ENDERECO,
                REGIAO_ADMINISTRATIVA,
                CONSENTIU_CONTATO,
                OBSERVACAO,
                DATA_CADASTRO,
                DATA_NASCIMENTO,
                CEP
            FROM CONTATOS_CAMPANHA
            WHERE CABO_ID = :1
            ORDER BY ID
        """, (cabo_id,))
        contatos = cursor.fetchall()

    except oracledb.Error as erro:
        flash(f"Erro ao exportar contatos: {erro}", "danger")
        print("ERRO AO EXPORTAR CONTATOS DO CABO:", erro)
        return redirect(url_for("listar"))
    finally:
        cursor.close()
        conexao.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contatos do Cabo"

    sheet.append([
        "ID",
        "Nome",
        "Telefone",
        "Endereço",
        "Região Administrativa",
        "Consentiu Contato",
        "Observação",
        "Data Cadastro",
        "Data Nascimento",
        "CEP"
    ])

    for contato in contatos:
        linha = list(contato)

        if linha[7]:
            linha[7] = linha[7].strftime("%d/%m/%Y")

        if linha[8]:
            linha[8] = linha[8].strftime("%d/%m/%Y")

        sheet.append(linha)

    arquivo_excel = BytesIO()
    workbook.save(arquivo_excel)
    arquivo_excel.seek(0)

    nome_arquivo = f"contatos_{nome_cabo.replace(' ', '_').lower()}.xlsx"

    return send_file(
        arquivo_excel,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/relatorio-geral")
@login_required
def relatorio_geral():
    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("home"))

    cursor = conexao.cursor()

    total_cabos = 0
    total_contatos = 0
    contatos_por_regiao = []
    ranking_cabos = []

    try:
        cursor.execute("SELECT COUNT(*) FROM CABOS_ELEITORAIS")
        total_cabos = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM CONTATOS_CAMPANHA")
        total_contatos = cursor.fetchone()[0]

        cursor.execute("""
            SELECT
                NVL(REGIAO_ADMINISTRATIVA, 'Não informada') AS REGIAO,
                COUNT(*) AS TOTAL
            FROM CONTATOS_CAMPANHA
            GROUP BY NVL(REGIAO_ADMINISTRATIVA, 'Não informada')
            ORDER BY TOTAL DESC, REGIAO
        """)
        contatos_por_regiao = cursor.fetchall()

        cursor.execute("""
            SELECT
                c.NOME,
                NVL(c.REGIAO_ADMINISTRATIVA, 'Não informada') AS REGIAO,
                COUNT(ct.ID) AS TOTAL_CONTATOS
            FROM CABOS_ELEITORAIS c
            LEFT JOIN CONTATOS_CAMPANHA ct ON ct.CABO_ID = c.ID
            GROUP BY c.NOME, NVL(c.REGIAO_ADMINISTRATIVA, 'Não informada')
            ORDER BY TOTAL_CONTATOS DESC, c.NOME
        """)
        ranking_cabos = cursor.fetchall()

    except oracledb.Error as erro:
        flash(f"Erro ao carregar relatório: {erro}", "danger")
        return redirect(url_for("home"))
    finally:
        cursor.close()
        conexao.close()

    regioes_labels = [item[0] for item in contatos_por_regiao]
    regioes_valores = [item[1] for item in contatos_por_regiao]

    cabos_labels = [item[0] for item in ranking_cabos]
    cabos_valores = [item[2] for item in ranking_cabos]

    return render_template(
        "relatorio_geral.html",
        total_cabos=total_cabos,
        total_contatos=total_contatos,
        contatos_por_regiao=contatos_por_regiao,
        ranking_cabos=ranking_cabos,
        regioes_labels=regioes_labels,
        regioes_valores=regioes_valores,
        cabos_labels=cabos_labels,
        cabos_valores=cabos_valores
    )

@app.route("/exportar-relatorio-geral-excel")
def exportar_relatorio_geral_excel():
    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("relatorio_geral"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT
                c.NOME,
                NVL(c.REGIAO_ADMINISTRATIVA, 'Não informada') AS REGIAO,
                COUNT(ct.ID) AS TOTAL_CONTATOS
            FROM CABOS_ELEITORAIS c
            LEFT JOIN CONTATOS_CAMPANHA ct ON ct.CABO_ID = c.ID
            GROUP BY c.NOME, NVL(c.REGIAO_ADMINISTRATIVA, 'Não informada')
            ORDER BY TOTAL_CONTATOS DESC, c.NOME
        """)
        ranking_cabos = cursor.fetchall()

        cursor.execute("""
            SELECT
                NVL(REGIAO_ADMINISTRATIVA, 'Não informada') AS REGIAO,
                COUNT(*) AS TOTAL
            FROM CONTATOS_CAMPANHA
            GROUP BY NVL(REGIAO_ADMINISTRATIVA, 'Não informada')
            ORDER BY TOTAL DESC, REGIAO
        """)
        contatos_por_regiao = cursor.fetchall()

    except oracledb.Error as erro:
        flash(f"Erro ao exportar relatório: {erro}", "danger")
        return redirect(url_for("relatorio_geral"))
    finally:
        cursor.close()
        conexao.close()

    workbook = Workbook()

    aba1 = workbook.active
    aba1.title = "Ranking Cabos"
    aba1.append(["Cabo", "Região", "Total de Contatos"])
    for linha in ranking_cabos:
        aba1.append(list(linha))

    aba2 = workbook.create_sheet(title="Contatos por Região")
    aba2.append(["Região", "Total de Contatos"])
    for linha in contatos_por_regiao:
        aba2.append(list(linha))

    arquivo_excel = BytesIO()
    workbook.save(arquivo_excel)
    arquivo_excel.seek(0)

    return send_file(
        arquivo_excel,
        as_attachment=True,
        download_name="relatorio_geral.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        senha = request.form.get("senha", "").strip()

        if usuario == ADMIN_USER and senha == ADMIN_PASSWORD:
            session["usuario_logado"] = usuario
            flash("Login realizado com sucesso.", "success")
            return redirect(url_for("home"))
        else:
            flash("Usuário ou senha inválidos.", "danger")

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("Logout realizado com sucesso.", "success")
    return redirect(url_for("login"))

@app.route("/enviar-convite/<int:cabo_id>", methods=["GET", "POST"])
@login_required
def enviar_convite(cabo_id):
    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("listar"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT ID, NOME, EMAIL
            FROM CABOS_ELEITORAIS
            WHERE ID = :1
        """, (cabo_id,))
        cabo = cursor.fetchone()

        if not cabo:
            flash("Liderança não encontrada.", "warning")
            return redirect(url_for("listar"))

        if request.method == "POST":
            email = request.form.get("email", "").strip().lower()

            if not email:
                flash("Informe o e-mail do contato.", "warning")
                return render_template("enviar_convite.html", cabo=cabo)

            cursor.execute("""
                SELECT ID
                FROM CONVITES_CONTATO
                WHERE EMAIL = :1
                  AND CABO_ID = :2
                  AND STATUS = 'PENDENTE'
            """, (email, cabo_id))
            convite_existente = cursor.fetchone()

            if convite_existente:
                flash("Já existe um convite pendente para este e-mail.", "warning")
                return render_template("enviar_convite.html", cabo=cabo)

            token = gerar_token_convite()

            cursor.execute("""
                INSERT INTO CONVITES_CONTATO (EMAIL, CABO_ID, TOKEN, STATUS, DATA_EXPIRACAO)
                VALUES (:1, :2, :3, 'PENDENTE', SYSDATE + 7)
            """, (email, cabo_id, token))
            conexao.commit()

            link_convite = f"{BASE_URL}/convite/{token}"

            enviar_email_convite(email, cabo[1], link_convite)

            flash("Convite enviado com sucesso.", "success")
            return redirect(url_for("listar"))

    except oracledb.Error as erro:
        flash(f"Erro ao processar convite: {erro}", "danger")
        print("ERRO AO ENVIAR CONVITE:", erro)
        return redirect(url_for("listar"))
    except Exception as erro:
        flash(f"Erro ao enviar e-mail: {erro}", "danger")
        print("ERRO SMTP:", erro)
        return redirect(url_for("listar"))
    finally:
        cursor.close()
        conexao.close()

    return render_template("enviar_convite.html", cabo=cabo)

@app.route("/convite/<token>", methods=["GET", "POST"])
def cadastro_por_convite(token):
    conexao = conectar_oracle()
    if conexao is None:
        flash("Não foi possível conectar ao banco de dados.", "danger")
        return redirect(url_for("login"))

    cursor = conexao.cursor()

    try:
        cursor.execute("""
            SELECT ID, EMAIL, CABO_ID, STATUS, DATA_EXPIRACAO
            FROM CONVITES_CONTATO
            WHERE TOKEN = :1
        """, (token,))
        convite = cursor.fetchone()

        if not convite:
            flash("Convite inválido.", "danger")
            return redirect(url_for("login"))

        convite_id = convite[0]
        email = convite[1]
        cabo_id = convite[2]
        status = convite[3]
        data_expiracao = convite[4]

        if data_expiracao and data_expiracao < datetime.now():
            cursor.execute("""
                UPDATE CONVITES_CONTATO
                SET STATUS = 'EXPIRADO'
                WHERE ID = :1
                  AND STATUS = 'PENDENTE'
            """, (convite_id,))
            conexao.commit()
            flash("Este convite está expirado.", "warning")
            return redirect(url_for("login"))

        if status != "PENDENTE":
            flash("Este convite já foi utilizado ou não está mais disponível.", "warning")
            return redirect(url_for("login"))

        cursor.execute("""
            SELECT NOME
            FROM CABOS_ELEITORAIS
            WHERE ID = :1
        """, (cabo_id,))
        cabo = cursor.fetchone()

        if not cabo:
            flash("Liderança vinculada ao convite não encontrada.", "danger")
            return redirect(url_for("login"))

        nome_cabo = cabo[0]

        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            telefone = request.form.get("telefone", "").strip()
            endereco = request.form.get("endereco", "").strip()
            regiao = request.form.get("regiao_administrativa", "").strip()
            data_nascimento = converter_data(request.form.get("data_nascimento", "").strip())
            cep = request.form.get("cep", "").strip()
            consentiu = request.form.get("consentiu_contato", "N")
            observacao = request.form.get("observacao", "").strip()

            if not nome:
                flash("O nome é obrigatório.", "warning")
                return render_template(
                    "cadastro_por_convite.html",
                    email=email,
                    nome_cabo=nome_cabo
                )

            cursor.execute("""
                SELECT ID
                FROM CONTATOS_CAMPANHA
                WHERE UPPER(NOME) = UPPER(:1)
                  AND CABO_ID = :2
            """, (nome, cabo_id))
            registro_existente = cursor.fetchone()

            if registro_existente:
                flash("Já existe um contato cadastrado para esta liderança com esse nome.", "warning")
                return render_template(
                    "cadastro_por_convite.html",
                    email=email,
                    nome_cabo=nome_cabo
                )

            cursor.execute("""
                INSERT INTO CONTATOS_CAMPANHA
                (NOME, TELEFONE, ENDERECO, REGIAO_ADMINISTRATIVA, CABO_ID,
                 CONSENTIU_CONTATO, OBSERVACAO, DATA_NASCIMENTO, CEP)
                VALUES (:1, :2, :3, :4, :5, :6, :7, :8, :9)
            """, (
                nome,
                telefone,
                endereco,
                regiao,
                cabo_id,
                consentiu,
                observacao,
                data_nascimento,
                cep
            ))

            cursor.execute("""
                UPDATE CONVITES_CONTATO
                SET STATUS = 'USADO',
                    DATA_USO = SYSDATE
                WHERE ID = :1
            """, (convite_id,))

            conexao.commit()
            flash("Cadastro concluído com sucesso.", "success")
            return redirect(url_for("login"))

    except oracledb.Error as erro:
        flash(f"Erro ao processar cadastro por convite: {erro}", "danger")
        print("ERRO CADASTRO POR CONVITE:", erro)
        return redirect(url_for("login"))
    finally:
        cursor.close()
        conexao.close()

    return render_template(
        "cadastro_por_convite.html",
        email=email,
        nome_cabo=nome_cabo
    )

#Criar funçoes para converter data e validar campos de formulário, se necessário.

def gerar_token_convite():
    return secrets.token_urlsafe(32)


def enviar_email_convite(destinatario, nome_cabo, link):
    assunto = "Convite para concluir seu cadastro"
    corpo_html = f"""
    <html>
        <body>
            <h3>Olá!</h3>
            <p>Você recebeu um convite para concluir seu cadastro.</p>
            <p><strong>Responsável pelo convite:</strong> {nome_cabo}</p>
            <p>Clique no link abaixo para continuar:</p>
            <p>
                <a href="{link}" style="background:#28a745;color:#fff;padding:10px 16px;text-decoration:none;border-radius:6px;">
                    Concluir cadastro
                </a>
            </p>
            <p>Ou copie e cole este endereço no navegador:</p>
            <p>{link}</p>
            <p>Se você não solicitou isso, ignore este e-mail.</p>
        </body>
    </html>
    """

    msg = MIMEMultipart("alternative")
    msg["Subject"] = assunto
    msg["From"] = MAIL_FROM
    msg["To"] = destinatario

    msg.attach(MIMEText(corpo_html, "html", "utf-8"))

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as servidor:
        servidor.starttls()
        servidor.login(SMTP_USER, SMTP_PASSWORD)
        servidor.sendmail(MAIL_FROM, destinatario, msg.as_string())

if __name__ == "__main__":
    debug = os.getenv("FLASK_DEBUG", "False").lower() == "true"
    app.run(debug=debug)
    
    
  



